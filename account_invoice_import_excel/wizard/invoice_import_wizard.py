from odoo import models, fields, api, _
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import requests
import xml.etree.ElementTree as ET
import xlwt
import io
import re
import base64
import openpyxl
from odoo.exceptions import UserError
from datetime import datetime, date

class InvoiceImportWizard(models.TransientModel):
    _name = "invoice.import.wizard"
    _description = "Import Invoices from Excel"

    journal_id = fields.Many2one(
        'account.journal',
        string="ჟურნალი",
        required=True,
        domain=[('type', '=', 'sale')],
    )

    line_from = fields.Integer(string='პირველი ხაზი')
    line_to = fields.Integer(string='ბოლო ხაზი')
    invoice_date = fields.Date(string='ინვოისის თარიღი')

    file = fields.Binary("File", required=True)
    filename = fields.Char("Filename")
    def _validate_vat_on_rs(self, vat):
        if not vat:
            return False

        vat = str(vat).strip().replace('\xa0', '')
        if vat.endswith('.0'):
            vat = vat[:-2]

        if not re.match(r'^(\d{9}|\d{11})$', vat):
            return False

        # 2. Prepare API Credentials
        usn = self.env.user.rs_acc
        usp = self.env.user.rs_pass
        if not usn or not usp:
            return False

        soap_request = f"""<?xml version="1.0" encoding="utf-8"?>
        <soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
          <soap:Body>
            <get_name_from_tin xmlns="http://tempuri.org/">
              <su>{usn}</su><sp>{usp}</sp><tin>{vat}</tin>
            </get_name_from_tin>
          </soap:Body>
        </soap:Envelope>"""

        headers = {
            "Content-Type": "text/xml; charset=utf-8",
            "SOAPAction": "http://tempuri.org/get_name_from_tin"
        }

        try:
            response = requests.post(
                "http://services.rs.ge/waybillservice/waybillservice.asmx",
                data=soap_request,
                headers=headers,
                timeout=5
            )

            if response.status_code == 200:
                root = ET.fromstring(response.content)
                result_node = root.find('.//{http://tempuri.org/}get_name_from_tinResult')

                if result_node is None:
                    return False

                name_text = result_node.text

                if not name_text:
                    return False

                name_text = name_text.strip()
                if not name_text:
                    return False

                if name_text.lower() == 'null':
                    return False

                return True

        except Exception as e:
            _logger.error(f"RS Validation Failed for {vat}: {e}")
            return False

        return False

    def _export_error_report(self, failed_lines):
        workbook = xlwt.Workbook(encoding='utf-8')
        sheet = workbook.add_sheet('Failed Imports')
        headers = ['VAT Code', 'Partner in Excel', 'Reason']
        for i, h in enumerate(headers): sheet.write(0, i, h)
        for row_no, data in enumerate(failed_lines, start=1):
            sheet.write(row_no, 0, data['vat'])
            sheet.write(row_no, 1, data['excel_name'])
            sheet.write(row_no, 2, data['reason'])
        fp = io.BytesIO()
        workbook.save(fp)
        export_id = self.env['ir.attachment'].create({
            'name': 'failed_partners_report.xls',
            'type': 'binary',
            'datas': base64.b64encode(fp.getvalue()),
            'mimetype': 'application/vnd.ms-excel'
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{export_id.id}?download=true', 'target': 'self'}


    def _sanitize_vat(self, value):
        if not value:
            return False
        val_str = str(value).strip()
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        return val_str

    def action_import_invoices(self):
        if not self.file:
            raise UserError(_("გთხოვთ ატვირთოთ ფაილი სანამ იმპორტს დაიწყებთ."))

        try:
            file_content = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(("Excel ფაილის წაკითხვა ვერ მოხერხდა: %s") % str(e))

        invoices = self.env['account.move']
        statement_lines = self.env['account.bank.statement.line']

        def parse_number(value):
            if value is None:
                return 0.0
            try:
                return float(str(value).replace(',', '').strip())
            except Exception:
                return 0.0

        journal_name = self.journal_id.name.strip() if self.journal_id else ""
        Move = self.env['account.move']
        Partner = self.env['res.partner']  # Define Partner model shortcut
        invoice_year = self.invoice_date.year if self.invoice_date else fields.Date.today().year
        failed_lines = []
        # ───────────────────────────────
        # სარეკლამო მომსახურება
        # ───────────────────────────────
        try:
            if "სარეკლამო მომსახურება" in journal_name:
                invoice_counters = {}
                Product = self.env['product.product']
                ids = Product.search_read([('name', '=', 'სარეკლამო მომსახურება')], fields=['id'], limit=1)
                product = Product.browse(ids[0]['id']) if ids else False
                if not product:
                    product = Product.create({'name': 'სარეკლამო მომსახურება', 'type': 'service'})

                min_row = self.line_from or 4
                max_row = self.line_to or sheet.max_row

                for row_idx, row in enumerate(sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True),
                                              start=min_row):
                    partner_name = str(row[0]).strip() if row[0] else False
                    if not partner_name:
                        continue

                    # FIX: Sanitize VAT
                    identification_code = self._sanitize_vat(row[1])

                    amount = parse_number(row[9])

                    #Only search by VAT if VAT exists
                    partner = False
                    if identification_code:
                        partner = Partner.search([('vat', '=', identification_code)], limit=1)

                    if not partner:
                            # Create with sanitized VAT immediately
                        if self._validate_vat_on_rs(identification_code):
                            partner = Partner.create({
                                'name': partner_name,
                                'vat': identification_code,
                                'company_type': 'company'})
                        else:
                            failed_lines.append({
                                'vat': identification_code,
                                'excel_name': partner.name,
                                'reason': 'Invalid VAT on RS.GE'
                            })
                            continue

                    write_vals = {}
                    if identification_code and partner.vat != identification_code:
                        write_vals['vat'] = identification_code

                    # Determine company type
                    new_type = 'company'
                    if partner.company_type != new_type:
                        write_vals['company_type'] = new_type


                    if write_vals:
                        partner.write(write_vals)

                    # Ensure VAT is updated on existing partner if missing

                    comment = row[10]
                    prefix = 'SM'
                    invoice_name = self._get_next_invoice_name(prefix, invoice_year, Move, invoice_counters)

                    move_vals = {
                        'name': invoice_name,
                        'move_type': 'out_invoice',
                        'partner_id': partner.id,
                        'product_comment': comment,
                        'invoice_date': self.invoice_date,
                        'invoice_date_due': self.invoice_date,
                        'identification_code': identification_code or '',
                        'journal_id': self.journal_id.id,
                        'invoice_line_ids': [(0, 0, {
                            'product_id': product.id,
                            'name': 'სარეკლამო მომსახურება',
                            'quantity': 1,
                            'price_unit': amount,
                        })]
                    }
                    move = Move.create(move_vals)
                    invoices |= move

            # ───────────────────────────────
            # მიწისქვეშა გადასასვლელების მოიჯარეები
            # ───────────────────────────────
            elif any(keyword in journal_name for keyword in ["საიჯარო", "მოიჯარე", "მიწისქვეშა", "მოიჯარეები"]):
                invoice_counters = {}
                Product = self.env['product.product']
                ids = Product.search_read([('name', '=', 'იჯარა')], fields=['id'], limit=1)
                product = Product.browse(ids[0]['id']) if ids else False
                if not product:
                    product = self.env['product.product'].create({'name': 'იჯარა', 'type': 'service'})

                min_row = self.line_from or 2
                max_row = self.line_to if self.line_to else sheet.max_row

                for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
                    contract_num = row[1]
                    basis = row[6]
                    partner_name = str(row[3]).strip() if row[3] else False

                    # Sanitize VAT
                    identification_code = self._sanitize_vat(row[4])
                    monthly_amount = parse_number(row[5])

                    if not partner_name:
                        continue

                    # Safe Search
                    partner = False
                    if identification_code:
                        partner = Partner.search([('vat', '=', identification_code)], limit=1)


                    if not partner:
                        if self._validate_vat_on_rs(identification_code):
                            partner = Partner.create({
                                'name': partner_name,
                                'vat': identification_code,
                                'company_type': 'company' })
                        else:
                            failed_lines.append({
                                'vat': identification_code,
                                'excel_name': partner.name,
                                'reason': 'Invalid VAT on RS.GE'
                            })
                            continue
                    write_vals = {}

                    if identification_code and partner.vat != identification_code:
                        write_vals['vat'] = identification_code

                    # Determine company type
                    new_type = 'company'
                    if partner.company_type != new_type:
                        write_vals['company_type'] = new_type


                    if write_vals:
                        partner.write(write_vals)


                    basis_text = f"{basis or ''}".strip()
                    prefix = 'IJR'

                    invoice_name = self._get_next_invoice_name(prefix, invoice_year, Move, invoice_counters)
                    comment = row[6]

                    move = self.env['account.move'].create({
                        'name': invoice_name,
                        'move_type': 'out_invoice',
                        'partner_id': partner.id,
                        'invoice_date': self.invoice_date,
                        'product_comment': comment,
                        'invoice_date_due': self.invoice_date,
                        'identification_code': identification_code or '',
                        'journal_id': self.journal_id.id,
                        'contract_num': contract_num or '',
                        'basis': basis_text,
                        'invoice_line_ids': [(0, 0, {
                            'product_id': product.id,
                            'name': 'იჯარა',
                            'quantity': 1,
                            'price_unit': monthly_amount,
                        })]
                    })
                    invoices |= move

            # ───────────────────────────────
            # სხვა / სამშენებლო
            # ───────────────────────────────
            else:
                journal = self.env['account.journal'].search([('name', '=', journal_name)], limit=1)

                # Logic check for construction
                if ("სამშენებლო" in journal_name or "ნარჩენ" in journal_name) and (
                        ("(ფიზიკური " in journal_name) or ("(იურიდიული " in journal_name)):

                    if "(ფიზიკური " in journal_name:
                        partner_tag_id = self.env['res.partner.category'].search([('name', '=', 'ფიზიკური')], limit=1)
                    else:
                        partner_tag_id = self.env['res.partner.category'].search([('name', '=', 'იურიდიული')], limit=1)

                    Product = self.env['product.product']
                    ids = Product.search_read([('name', '=', 'სამშენებლო მასალები')], fields=['id'], limit=1)
                    product = Product.browse(ids[0]['id']) if ids else False
                    if not product:
                        raise UserError(_("Product 'სამშენებლო მასალები' not found"))

                    # Find header row
                    header_row_index = None
                    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        # Ensure we convert cell to string to check 'in' safely
                        row_str = [str(c).strip() if c else '' for c in row]
                        if 'კონტრაქტის ნომერი' in row_str:
                            header_row_index = i
                            break

                    if header_row_index is None:
                        raise UserError(_("Couldn't find header row (expected a row with 'კონტრაქტი')"))

                    headers = [str(cell).strip() if cell else '' for cell in
                               next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index,
                                                    values_only=True))]
                    col_index = {header: idx for idx, header in enumerate(headers)}

                    invoice_counters = {}
                    moves_buffer = []
                    invoices = self.env['account.move']
                    batch_size = 200

                    # Contexts
                    ctx = dict(self.env.context or {}, mail_notrack=True, tracking_disable=True)
                    Move = self.env['account.move'].with_context(ctx)
                    Partner = self.env['res.partner'].with_context(ctx)
                    created_ids = []

                    min_row = self.line_from or (header_row_index + 1)
                    max_row = self.line_to if self.line_to else sheet.max_row

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True),
                                                  start=min_row):
                        if not row or not row[col_index.get('კლიენტი')]:
                            continue

                        # Sanitize VAT here specifically
                        identification_num = self._sanitize_vat(row[col_index['საიდენტიფიკაციო']])
                        partner_name = str(row[col_index['კლიენტი']]).strip()

                        # Safe Search
                        partner = False
                        if identification_num:
                            partner = Partner.search([('vat', '=', identification_num)], limit=1)

                        if not partner:
                            if self._validate_vat_on_rs(identification_num):
                                partner = Partner.create({
                                    'name': partner_name,
                                    'vat': identification_num,
                                    'category_id': partner_tag_id
                                })
                            else:
                                failed_lines.append({
                                    'vat': identification_num,
                                    'excel_name': partner_name,
                                    'reason': 'Invalid VAT on RS.GE'
                                })
                                continue


                        write_vals = {}
                        if identification_num and partner.vat != identification_num:
                            write_vals['vat'] = identification_num

                        # Determine company type
                        c_type = row[col_index.get('ტიპი')]
                        new_type = 'company' if c_type and str(c_type).strip() in ['შ.პ.ს', 'ი/მ', 'ს/ს'] else 'person'
                        if partner.company_type != new_type:
                            write_vals['company_type'] = new_type

                        if partner.category_id != partner_tag_id :
                            write_vals['category_id'] = partner_tag_id

                        if write_vals:
                            partner.write(write_vals)

                        partner_id = partner.id
                        parsed_date = self.parse_excel_date(row[col_index['თარიღი']])
                        invoice_year = parsed_date.year if parsed_date else datetime.today().year

                        prefix = 'SNP' if '(ფიზიკური' in journal.name else 'SNI'
                        counter_key = (prefix, invoice_year)

                        # Counter Logic
                        if counter_key not in invoice_counters:
                            last_invoice = Move.search([
                                ('move_type', '=', 'out_invoice'),
                                ('name', 'like', f'{prefix}/{invoice_year}/%')
                            ], order='id desc', limit=1)
                            if last_invoice and last_invoice.name:
                                try:
                                    last_index = int(last_invoice.name.split('/')[-1])
                                except Exception:
                                    last_index = 0
                            else:
                                last_index = 0
                            invoice_counters[counter_key] = last_index

                        invoice_counters[counter_key] += 1
                        new_index = invoice_counters[counter_key]
                        invoice_name = f"{prefix}/{invoice_year}/{str(new_index).zfill(3)}"

                        move_vals = {
                            'name': invoice_name,
                            'move_type': 'out_invoice',
                            'partner_id': partner_id,
                            'journal_id': journal.id,
                            'invoice_date': parsed_date,
                            'invoice_date_due': parsed_date,
                            'product_comment': row[col_index['ტექსტი']],
                            'identification_code': identification_num,
                            'contract_num': row[col_index['კონტრაქტის ნომერი']],
                            'invoice_line_ids': [(0, 0, {
                                'product_id': product.id,
                                'quantity': 1.0,
                                'price_unit': row[col_index['სრული ფასი']],
                            })]
                        }
                        if '(იურიდიული ' in journal.name:
                            try:
                                factura_num = row[col_index['ა/ფ']]
                                move_vals['get_invoice_id_helper'] = factura_num
                            except Exception:
                                move_vals['get_invoice_id_helper'] = ''

                        moves_buffer.append(move_vals)

                        if len(moves_buffer) >= batch_size:
                            created = Move.create(moves_buffer)
                            self.env.cr.commit()
                            created_ids.extend(created.ids)
                            moves_buffer.clear()

                    if moves_buffer:
                        created = Move.create(moves_buffer)
                        created_ids.extend(created.ids)
                        self.env.cr.commit()

                    invoices = Move.browse(created_ids)

            if failed_lines:
                return self._export_error_report(failed_lines)

            if not invoices:
                raise UserError(_("არ შეიქმნა არცერთი ინვოისი. გთხოვთ გადაამოწმოთ ფაილი."))

        except UserError:
            raise

        except Exception as e:
            import traceback
            traceback.print_exc()
            raise UserError(_("ინვოისების იმპორტისას მოხდა შეცდომა: %s") % str(e))

        return {
            "type": "ir.actions.act_window",
            "name": "Invoices",
            "res_model": "account.move",
            "view_mode": "list,form",
            "domain": [("id", "in", invoices.ids)],
            "context": {'default_move_type': 'out_invoice'}
        }

    def _get_next_invoice_name(self, prefix, year, Move, counters):
        key = (prefix, year)
        if key not in counters:
            last = Move.search([
                ('move_type', '=', 'out_invoice'),
                ('name', 'like', f'{prefix}/{year}/%')
            ], order='id desc', limit=1)
            last_index = 0
            if last and last.name:
                try:
                    last_index = int(last.name.split('/')[-1])
                except Exception:
                    pass
            counters[key] = last_index
        counters[key] += 1
        return f"{prefix}/{year}/{str(counters[key]).zfill(3)}"

    @staticmethod
    def parse_excel_date(value):
        if not value:
            return False

        if isinstance(value, (datetime, date)):
            return value.date() if isinstance(value, datetime) else value

        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return False

        if isinstance(value, str):
            value = " ".join(value.strip().split())
            for fmt in (
                    "%d-%m-%Y %H:%M:%S",
                    "%Y-%m-%d %H:%M:%S",
                    "%d/%m/%Y %H:%M:%S",
                    "%Y/%m/%d %H:%M:%S",
                    "%d-%m-%Y",
                    "%Y-%m-%d",
                    "%d/%m/%Y",
                    "%m/%d/%Y",
                    "%d.%m.%Y",
                    "%d.%m.%Y %H:%M:%S",
            ):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue

        return False