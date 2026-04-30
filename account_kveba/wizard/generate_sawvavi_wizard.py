from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
import openpyxl


class GenerateSawvaviWizard(models.TransientModel):
    _name = 'generate.sawvavi.wizard'
    _description = 'Generate Sawvavi Wizard'

    date = fields.Date('Date', required=True)
    excel_file = fields.Binary('Excel File', required=True)
    excel_filename = fields.Char('Excel File Name', required=True)
    fuel_type = fields.Selection(
        selection=[
            ('1', 'ბენზინი'),
            ('2', 'დიზელი'),
        ],
        string='ფუელის ტიპი',
        required=True,
        default='1',
    )


    def _generate_and_download_missed_excel(self, missed_rows, created_count):
        try:
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'დაკარგული ჩანაწერები'
            
            # Header row
            headers = ['რიგი', 'თანამშრომლის სახელი', 'პირადი ნომერი', 'რაოდენობა']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Add missed rows
            for mr in missed_rows:
                ws.append([
                    mr.get('row'),
                    mr.get('employee_name'),
                    mr.get('identification_id'),
                    mr.get('quantity')
                ])
            
            # Save to binary
            stream = BytesIO()
            wb.save(stream)
            excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')
            
            # Generate filename
            filename = 'დაკარგული_ჩანაწერები_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Show message and download
            if created_count == 0:
                message = 'არცერთი ჩანაწერი არ შეიქმნა!\n%s ჩანაწერი ვერ დამუშავდა და ყველა ჩანაწერი უარყოფილია.' % len(missed_rows)
                title = 'ატვირთვა ვერ მოხერხდა'
                msg_type = 'danger'
            else:
                message = '%s ჩანაწერი შეიქმნა წარმატებით.\n%s ჩანაწერი ვერ დამუშავდა.' % (
                    created_count, len(missed_rows)
                )
                title = 'ატვირთვა დასრულდა'
                msg_type = 'warning'
            
            # Create temporary attachment for download
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': excel_data,
                'res_model': 'generate.kveba.wizard',
                'res_id': 0,
                'type': 'binary',
            })
            
            # Return download action with message
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': title,
                    'message': message,
                    'type': msg_type,
                    'sticky': False,
                    'next': {
                        'type': 'ir.actions.act_url',
                        'url': '/web/content/%s?download=true' % attachment.id,
                        'target': 'self',
                    }
                }
            }
            
        except Exception as e:
            raise UserError(_('ვერ შეიქმნა შეცდომების ფაილი: %s') % str(e))
       

    def generate_sawvavi(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Excel file is missing"))
        
        try:
            file_data = base64.b64decode(self.excel_file)
            file_stream = BytesIO(file_data)
        except Exception as e:
            raise UserError(_("Error decoding Excel file: %s") % str(e))

        try:
            workbook = openpyxl.load_workbook(file_stream, data_only=True)
            sheet = workbook.active
            max_row = sheet.max_row
        except Exception as e:
            raise UserError(_("Error loading Excel file: %s") % str(e))

        sawvavi = False
        if self.fuel_type == '1':
            sawvavi = self.env['product.template'].sudo().search([('display_name', '=', '[13395] პრემიუმ ავანგარდი')], limit=1)
        else:
            sawvavi = self.env['product.template'].sudo().search([('display_name', '=', '[11818] დიზელი')], limit=1)
        if not sawvavi:
            raise UserError(_("Product is not found"))
        
        account_1619 = self.env['account.account'].sudo().search([('code', '=', '1619')], limit=1)
        account_3135 = self.env['account.account'].sudo().search([('code', '=', '3135')], limit=1)
        account_7411 = self.env['account.account'].sudo().search([('code', '=', '7411')], limit=1)
        account_3322 = self.env['account.account'].sudo().search([('code', '=', '3322')], limit=1)
        account_3330 = self.env['account.account'].sudo().search([('code', '=', '3330')], limit=1)
        account_7410 = self.env['account.account'].sudo().search([('code', '=', '7410')], limit=1)
        account_3133 = self.env['account.account'].sudo().search([('code', '=', '3133')], limit=1)

        if not all([account_1619, account_3135, account_7411, account_3322, account_3330, account_7410, account_3133]):
            raise UserError(_("Required accounts not found. Please check account codes: 1619, 3132, 7411, 3322, 3330, 7410, 3133"))

        cost = sawvavi.standard_price
        
        invoice_lines = []
        missed_rows = []
        
        for row in range(2, max_row + 1):
            cell_value = sheet[f'M{row}'].value
            cell_value_2 = sheet[f'F{row}'].value
            cell_value_3 = sheet[f'L{row}'].value
            cell_value_4 = sheet[f'D{row}'].value
            cell_value_5 = sheet[f'G{row}'].value

            employee_name = str(cell_value_3) if cell_value_3 else ''
            car_number = str(cell_value_4) if cell_value_4 else ''
            analytic_account_name = str(cell_value_5) if cell_value_5 else ''
            
            analytic_account = False
            if analytic_account_name:
                analytic_account = self.env['account.analytic.account'].sudo().search([('name', '=', analytic_account_name)], limit=1)

            if not cell_value:
                continue
            if not cell_value_2:
                quantity = 0.0
            else:
                quantity = float(cell_value_2)
            
            identification_id = str(cell_value)
            if '.' in identification_id:
                identification_id = identification_id.split('.')[0]

            employee = False
            partner = False
            if identification_id:
                employee = self.env['hr.employee'].sudo().search([('identification_id', '=', identification_id)], limit=1)
                partner = employee.work_contact_id if employee else False
                if not partner:
                    missed_rows.append({
                        'row': row,
                        'employee_name': employee_name,
                        'identification_id': identification_id,
                        "quantity": quantity,
                    })
                    continue
            
            base_amount = quantity * cost
            has_pension = hasattr(partner, 'x_studio_') and partner.x_studio_

            # Line 1: (3132 - 1619) - საწვავი
            invoice_lines.append((0,0,  {
                'account_id': account_3135.id,
                'partner_id': partner.id,
                'name': f'3132 -{partner.name}',
                'quantity': quantity,
                'product_id': sawvavi.product_variant_id.id,
                'price_unit': cost,
                'debit': base_amount,
                'credit': 0.0,
            }))
            
            invoice_lines.append((0,0, {
                'account_id': account_1619.id,
                'partner_id': partner.id,
                'name': f'1618 -{partner.name}',
                'quantity': quantity,
                'product_id': sawvavi.product_variant_id.id,
                'price_unit': cost,
                'debit': 0.0,
                'credit': base_amount,
                'analytic_distribution': {str(analytic_account.id): 100.0} if analytic_account else False,
            }))

            if has_pension:
                amount_7411 = base_amount * 1.505102041
                amount_3322 = base_amount * 0.295
                amount_3330 = base_amount * 0.18
                amount_pension = base_amount * 0.030102041

                # Line 2: 7411 - 3132
                invoice_lines.append((0, 0, {
                    'account_id': account_7411.id,
                    'name': f'7411 - {partner.name}',
                    'debit': amount_7411,
                    'credit': 0.0,
                    'analytic_distribution': {str(analytic_account.id): 100.0} if analytic_account else False,
                }))

                invoice_lines.append((0, 0, {
                    'account_id': account_3135.id,
                    'partner_id': partner.id,
                    'name': f'3132 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_7411,
                }))

                # Line 3: 3132 - 3322
                invoice_lines.append((0, 0, {
                    'account_id': account_3135.id,
                    'partner_id': partner.id,
                    'name': f'3132 - {partner.name}',
                    'debit': amount_3322,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3322.id,
                    'name': f'3322 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_3322,
                }))

                # Line 4: 3132 - 3330
                invoice_lines.append((0, 0, {
                    'account_id': account_3135.id,
                    'partner_id': partner.id,
                    'name': f'3132 - {partner.name}',
                    'debit': amount_3330,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3330.id,
                    'name': f'3330 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_3330,
                }))

                # Line 5: 7410 - 3133 
                invoice_lines.append((0, 0, {
                    'account_id': account_7410.id,
                    'name': f'პენსია - {partner.name}',
                    'debit': amount_pension,
                    'credit': 0.0,
                    'analytic_distribution': {str(analytic_account.id): 100.0} if analytic_account else False,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3133.id,
                    'partner_id': partner.id,
                    'name': f'პენსია - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_pension,
                }))

                # Line 6: 3132 - 3133 
                invoice_lines.append((0, 0, {
                    'account_id': account_3135.id,
                    'partner_id': partner.id,
                    'name': f'პენსია - {partner.name}',
                    'debit': amount_pension,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3133.id,
                    'partner_id': partner.id,
                    'name': f'პენსია - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_pension,
                }))
            else:
                amount_7411 = base_amount * 1.475
                amount_3322 = base_amount * 0.295
                amount_3330 = base_amount * 0.18

                # Line 2: 7411 - 3132
                invoice_lines.append((0, 0, {
                    'account_id': account_7411.id,
                    'name': f'7411 - {partner.name}',
                    'debit': amount_7411,
                    'credit': 0.0,
                    'analytic_distribution': {str(analytic_account.id): 100.0} if analytic_account else False,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3135.id,
                    'partner_id': partner.id,
                    'name': f'3132 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_7411,
                }))

                # Line 3: 3132 - 3322
                invoice_lines.append((0, 0, {
                    'account_id': account_3135.id,
                    'partner_id': partner.id,
                    'name': f'3132 - {partner.name}',
                    'debit': amount_3322,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3322.id,
                    'name': f'3322 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_3322,
                }))
                
                # Line 4: 3132 - 3330
                invoice_lines.append((0, 0, {
                    'account_id': account_3135.id,
                    'partner_id': partner.id,
                    'name': f'3132 - {partner.name}',
                    'debit': amount_3330,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3330.id,
                    'name': f'3330 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_3330,
                }))
        
        move = False
        if invoice_lines:
            journal = self.env['account.journal'].sudo().search([('name', '=', 'საწვავი')], limit=1)
            if not journal:
                raise UserError(_("Journal 'საწვავი' is not found"))
            
            move = self.env['account.move'].create({
                'move_type': 'entry',
                'journal_id': journal.id,
                'date': self.date,
                'line_ids': invoice_lines,
            })
            
            # Commit the transaction to ensure move is saved
            self.env.cr.commit()
        
        # If there are missed rows, generate Excel and download
        if missed_rows:
            # Calculate how many records were created (approximate)
            created_count = len(invoice_lines) // 10 if invoice_lines else 0
            return self._generate_and_download_missed_excel(missed_rows, created_count)
        
        # If no missed rows and move was created, open the move
        if move:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'account.move',
                'res_id': move.id,
                'view_mode': 'form',
                'target': 'current',
            }
        else:
            raise UserError(_("No valid data found in Excel file"))