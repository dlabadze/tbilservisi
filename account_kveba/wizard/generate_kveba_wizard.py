import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import openpyxl


class GenerateKvebaWizard(models.TransientModel):
    _name = 'generate.kveba.wizard'
    _description = 'Generate Kveba Wizard'

    date = fields.Date('Date', required=True)
    excel_file = fields.Binary('Excel File', required=True)
    excel_filename = fields.Char('Excel File Name', required=True)
    journals = fields.Selection(
        selection=[
            ('1', 'კვება'),
            ('2', 'საჩუქარი'),
        ],
        string='ჟურნალი',
        required=True,
        default='1',
    )
    analytic_account = fields.Many2one(
        'account.analytic.account',
        string='Analytic Account',
        required=True,
        domain=[('plan_id.name', '=', 'დეპარტამენტი')]
    )
    cost = fields.Float(string='ღირებულება დღგ-ს გარეშე', required=True)
    start_row = fields.Integer(string='Start Row', default=2)
    
    def _generate_and_download_missed_excel(self, missed_rows, created_count):
        """Generate Excel file with missed rows and return download action."""
        try:
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'დაკარგული ჩანაწერები'
            
            # Header row
            headers = ['რიგი', 'თანამშრომლის სახელი', 'პირადი ნომერი', 'რაოდენობა']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Add missed rows
            for mr in missed_rows:
                ws.append([
                    mr.get('row'),
                    mr.get('employee_name'),
                    mr.get('identification_id'),
                    mr.get('quantity')
                ])
            
            # Save to binary
            stream = BytesIO()
            wb.save(stream)
            excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')
            
            # Generate filename
            filename = 'დაკარგული_ჩანაწერები_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Show message and download
            if created_count == 0:
                message = 'არცერთი ჩანაწერი არ შეიქმნა!\n%s ჩანაწერი ვერ დამუშავდა და ყველა ჩანაწერი უარყოფილია.' % len(missed_rows)
                title = 'ატვირთვა ვერ მოხერხდა'
                msg_type = 'danger'
            else:
                message = '%s ჩანაწერი შეიქმნა წარმატებით.\n%s ჩანაწერი ვერ დამუშავდა.' % (
                    created_count, len(missed_rows)
                )
                title = 'ატვირთვა დასრულდა'
                msg_type = 'warning'
            
            # Create temporary attachment for download
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': excel_data,
                'res_model': 'generate.kveba.wizard',
                'res_id': 0,
                'type': 'binary',
            })
            
            # Return download action with message
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': title,
                    'message': message,
                    'type': msg_type,
                    'sticky': False,
                    'next': {
                        'type': 'ir.actions.act_url',
                        'url': '/web/content/%s?download=true' % attachment.id,
                        'target': 'self',
                    }
                }
            }
            
        except Exception as e:
            raise UserError(_('ვერ შეიქმნა შეცდომების ფაილი: %s') % str(e))
    
    def action_generate_kveba(self):
        self.ensure_one()

        if self.journals == '1':
            kveba = self.env['product.template'].sudo().search([('name', '=', 'კვება')], limit=1)
        else:
            kveba = self.env['product.template'].sudo().search([('name', '=', 'საჩუქარი')], limit=1)
        if not kveba:
            raise UserError(_("Product is not found"))
        if not self.excel_file:
            raise UserError(_("Excel file is missing"))
        
        # 1. base64 decode
        try:
            file_data = base64.b64decode(self.excel_file)
            file_stream = BytesIO(file_data)
        except Exception as e:
            raise UserError(_("Error decoding Excel file: %s") % str(e))

        # 2. load excel
        try:
            workbook = openpyxl.load_workbook(file_stream, data_only=True)
            sheet = workbook.active
            max_row = sheet.max_row
        except Exception as e:
            raise UserError(_("Error loading Excel file. Please make sure it's a valid Excel file (.xlsx): %s") % str(e))
        
        # Get accounts with sudo()
        account_3132 = self.env['account.account'].sudo().search([('code', '=', '3132')], limit=1)
        account_1618_1617 = False
        if self.journals == '1':
            account_1618_1617 = self.env['account.account'].sudo().search([('code', '=', '1618')], limit=1)
        else:
            account_1618_1617 = self.env['account.account'].sudo().search([('code', '=', '1617')], limit=1)

        account_7411 = self.env['account.account'].sudo().search([('code', '=', '7411')], limit=1)
        account_3322 = self.env['account.account'].sudo().search([('code', '=', '3322')], limit=1)
        account_3330_01 = self.env['account.account'].sudo().search([('code', '=', '3330')], limit=1)
        account_7410 = self.env['account.account'].sudo().search([('code', '=', '7410')], limit=1)
        account_3133 = self.env['account.account'].sudo().search([('code', '=', '3133')], limit=1)
        
        if not all([account_3132, account_1618_1617, account_7411, account_3322, account_3330_01]):
            raise UserError(_("Required accounts not found. Please check account codes: 3132, 1618, 7411, 3322, 3330"))
        
        # Get cost (თვითღირებულება) from product
        cost = kveba.standard_price
        
        # Prepare invoice lines
        invoice_lines = []
        missed_rows = []
        
        for row in range(self.start_row, max_row + 1):
            cell_value = sheet[f'D{row}'].value
            cell_value_2 = sheet[f'E{row}'].value
            cell_value_3 = sheet[f'B{row}'].value

            employee_name = str(cell_value_3)

            if not cell_value:
                continue
            if not cell_value_2:
                quantity = 0.0
            else:
                quantity = float(cell_value_2)

            identification_id = str(cell_value)
            if '.' in identification_id:
                identification_id = identification_id.split('.')[0]

            # Find employee and partner
            partner = False
            employee = False
            if identification_id:
                employee = self.env['hr.employee'].sudo().search(
                    [('identification_id', '=', identification_id)],
                    limit=1
                )
                partner = employee.work_contact_id if employee else False
                if not partner:
                    # Add to missed rows with employee_name and identification_id
                    missed_rows.append({
                        'row': row,
                        'employee_name': employee_name,
                        'identification_id': identification_id,
                        'quantity': quantity,
                    })
                    continue
            
            # Check if employee has pension (საპენსიო)
            has_pension = hasattr(partner, 'x_studio_') and partner.x_studio_
            # raise UserError(f"has_pension: {has_pension} employee: {employee} employee name: {employee.name}")
            
            # Base amount: quantity * cost (რაოდენობა გამრავლებული თვითღირებულებაზე)
            base_amount = quantity * cost
            
            # Line 1: 3132 - 1618 (კვება product line)
            invoice_lines.append((0, 0, {
                'account_id': account_3132.id,
                'partner_id': partner.id,
                'name': f'{partner.name}',
                'quantity': quantity,
                'product_id': kveba.product_variant_id.id,
                'price_unit': cost,
                'debit': base_amount,
                'credit': 0.0,
            }))
            
            invoice_lines.append((0, 0, {
                'account_id': account_1618_1617.id,
                'name': f'{partner.name}',
                'quantity': quantity,
                'product_id': kveba.product_variant_id.id,
                'price_unit': cost,
                'debit': 0.0,
                'credit': base_amount,
                'analytic_distribution': {str(self.analytic_account.id): 100.0},
            }))
            
            # Calculate amounts based on pension status
            if has_pension:
                # With Pension - based on provided formulas
                # 451.5306122 / 300 = 1.505102041
                amount_7411 = base_amount * 1.505102041
                # 88.5 / 300 = 0.295
                amount_3322 = base_amount * 0.295
                # 54 / 300 = 0.18
                amount_3330 = base_amount * 0.18
                # 9.030612245 / 300 = 0.030102041
                amount_pension = base_amount * 0.030102041
                
                # Line 2: 7411 - 3132
                invoice_lines.append((0, 0, {
                    'account_id': account_7411.id,
                    'name': f'7411 - {partner.name}',
                    'debit': amount_7411,
                    'credit': 0.0,
                    'analytic_distribution': {str(self.analytic_account.id): 100.0},
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3132.id,
                    'partner_id': partner.id,
                    'name': f'7411 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_7411,
                }))
                
                # Line 3: 3132 - 3322
                invoice_lines.append((0, 0, {
                    'account_id': account_3132.id,
                    'partner_id': partner.id,
                    'name': f'3322 - {partner.name}',
                    'debit': amount_3322,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3322.id,
                    'name': f'3322 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_3322,
                    'analytic_distribution': {str(self.analytic_account.id): 100.0},
                }))
                
                # Line 4: 3132 - 3330.01
                invoice_lines.append((0, 0, {
                    'account_id': account_3132.id,
                    'partner_id': partner.id,
                    'name': f'{partner.name}',
                    'debit': amount_3330,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3330_01.id,
                    'name': f'{partner.name}',
                    'debit': 0.0,
                    'credit': amount_3330,
                    'analytic_distribution': {str(self.analytic_account.id): 100.0},
                }))
                
                # Line 5: 7410 - 3133 პარტნერი (Pension line)
                if account_7410 and account_3133:
                    invoice_lines.append((0, 0, {
                        'account_id': account_7410.id,
                        'name': f'პენსია - {partner.name}',
                        'debit': amount_pension,
                        'credit': 0.0,
                        'analytic_distribution': {str(self.analytic_account.id): 100.0},
                    }))
                    
                    invoice_lines.append((0, 0, {
                        'account_id': account_3133.id,
                        'partner_id': partner.id,
                        'name': f'პენსია - {partner.name}',
                        'debit': 0.0,
                        'credit': amount_pension,
                        'analytic_distribution': {str(self.analytic_account.id): 100.0},
                    }))
                    
                    # Line 6: 3132 პარტნერი - 3133 პარტნერი (Pension line)
                    invoice_lines.append((0, 0, {
                        'account_id': account_3132.id,
                        'partner_id': partner.id,
                        'name': f'პენსია - {partner.name}',
                        'debit': amount_pension,
                        'credit': 0.0,
                    }))
                    
                    invoice_lines.append((0, 0, {
                        'account_id': account_3133.id,
                        'partner_id': partner.id,
                        'name': f'პენსია - {partner.name}',
                        'debit': 0.0,
                        'credit': amount_pension,
                        'analytic_distribution': {str(self.analytic_account.id): 100.0},
                    }))
            else:
                # Without Pension - based on provided formulas
                # 442.5 / 300 = 1.475
                amount_7411 = base_amount * 1.475
                # 88.5 / 300 = 0.295
                amount_3322 = base_amount * 0.295
                # 54 / 300 = 0.18
                amount_3330 = base_amount * 0.18
                
                # Line 2: 7411 - 3132
                invoice_lines.append((0, 0, {
                    'account_id': account_7411.id,
                    'name': f'7411 - {partner.name}',
                    'debit': amount_7411,
                    'credit': 0.0,
                    'analytic_distribution': {str(self.analytic_account.id): 100.0},
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3132.id,
                    'partner_id': partner.id,
                    'name': f'7411 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_7411,
                }))
                
                # Line 3: 3132 - 3322
                invoice_lines.append((0, 0, {
                    'account_id': account_3132.id,
                    'partner_id': partner.id,
                    'name': f'3322 - {partner.name}',
                    'debit': amount_3322,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3322.id,
                    'name': f'3322 - {partner.name}',
                    'debit': 0.0,
                    'credit': amount_3322,
                    'analytic_distribution': {str(self.analytic_account.id): 100.0},
                }))
                
                # Line 4: 3132 - 3330.01
                invoice_lines.append((0, 0, {
                    'account_id': account_3132.id,
                    'partner_id': partner.id,
                    'name': f'{partner.name}',
                    'debit': amount_3330,
                    'credit': 0.0,
                }))
                
                invoice_lines.append((0, 0, {
                    'account_id': account_3330_01.id,
                    'name': f'{partner.name}',
                    'debit': 0.0,
                    'credit': amount_3330,
                    'analytic_distribution': {str(self.analytic_account.id): 100.0},
                }))
        
        # Create account.move
        if invoice_lines:
            # Search for journal with name "კვება"
            # journal = False
            if self.journals == '1':
                journal_name = 'კვება'
            else:
                journal_name = 'საჩუქრები'

            journal = self.env['account.journal'].sudo().search([('name', '=', journal_name)], limit=1)

            if not journal:
                raise UserError(_("Journal '%s' not found") % journal_name)
            
            move = self.env['account.move'].sudo().create({
                'move_type': 'entry',
                'date': self.date,
                'journal_id': journal.id,
                'line_ids': invoice_lines,
            })
            
            # Generate and download missed rows Excel if there are any
            if missed_rows:
                created_count = len(invoice_lines) // 10  # Approximate number of records created
                return self._generate_and_download_missed_excel(missed_rows, created_count)
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'account.move',
                'res_id': move.id,
                'view_mode': 'form',
                'target': 'current',
            }
        else:
            raise UserError(_("No valid data found in Excel file"))

