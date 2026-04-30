from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

_logger = logging.getLogger(__name__)

class SalaryPaymentImport(models.Model):
    _name = 'salary.payment.import'
    _description = 'Salary Payment Import'
    _order = 'create_date desc'

    name = fields.Char(string='Reference', required=True, copy=False, default='New')
    date = fields.Date(string='Date', default=fields.Date.context_today, required=True)
    excel_file = fields.Binary(string='Excel File', required=True)
    excel_filename = fields.Char(string='File Name')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('imported', 'Imported'),
        ('posted', 'Posted')
    ], string='Status', default='draft', copy=False)
    line_ids = fields.One2many('salary.payment.import.line', 'import_id', string='Payment Lines')
    journal_entry_count = fields.Integer(compute='_compute_journal_entry_count')
    journal_entry_ids = fields.Many2many('account.move', string='Journal Entries', copy=False)
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    department_id = fields.Many2one('hr.department', string='Department')
    
    def _compute_journal_entry_count(self):
        for record in self:
            record.journal_entry_count = len(record.journal_entry_ids)
    
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'New') == 'New':
                vals['name'] = self.env['ir.sequence'].next_by_code('salary.payment.import') or 'New'
        return super().create(vals_list)
    
    def action_import_excel(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Please upload an Excel file."))
        
        if not HAS_PANDAS:
            raise UserError(_("Excel import feature is not available. Please contact your system administrator to install the required components."))
        
        try:
            excel_data = base64.b64decode(self.excel_file)
        except Exception as e:
            raise UserError(_("Failed to decode Excel file: %s") % str(e))
        
        # Read Excel file
        try:
            df = pd.read_excel(io.BytesIO(excel_data), header=None, dtype=str)
        except Exception as e:
            raise UserError(_("Could not read Excel file: %s") % str(e))
        
        if df is None or df.empty:
            raise UserError(_("Excel file is empty or could not be read."))
        
        # Clear existing lines
        self.line_ids.unlink()
        
        COLUMN_LABEL_ROW_INDEX = 3     # 4th row (1-based)
        ACCOUNT_CONFIG_ROW_INDEX = 5   # 6th row (1-based)
        DATA_START_ROW_INDEX = 6       # 7th row (1-based)
        PERSONAL_NUMBER_COL_INDEX = 2  # Column C
        AMOUNT_COLUMNS = list(range(13, 25))  # Columns N (13) through Y (24)
        
        if len(df) <= ACCOUNT_CONFIG_ROW_INDEX:
            raise UserError(_("Excel file must contain at least 6 rows for account configuration."))
        
        column_labels = {}
        account_pairs = {}
        for col_idx in AMOUNT_COLUMNS:
            label_value = ''
            if len(df) > COLUMN_LABEL_ROW_INDEX and df.shape[1] > col_idx:
                label_value = self._safe_string(df.iloc[COLUMN_LABEL_ROW_INDEX, col_idx])
            column_labels[col_idx] = label_value
            header_value = ''
            if df.shape[1] > col_idx:
                header_value = self._safe_string(df.iloc[ACCOUNT_CONFIG_ROW_INDEX, col_idx])
            if not header_value:
                continue
            normalized_value = header_value.replace('–', '-')
            parts = [part.strip() for part in normalized_value.split('-') if part.strip()]
            if len(parts) < 2:
                _logger.warning(f"Skipping column {self._column_letter(col_idx)} - invalid debit-credit format: {header_value}")
                continue
            debit_code, credit_code = parts[0], parts[1]
            # Ensure accounts exist (at least inform user if not)
            if not self._account_exists(debit_code):
                raise UserError(_("Debit account with code %s was not found for column %s.") % (debit_code, self._column_letter(col_idx)))
            if not self._account_exists(credit_code):
                raise UserError(_("Credit account with code %s was not found for column %s.") % (credit_code, self._column_letter(col_idx)))
            account_pairs[col_idx] = (debit_code, credit_code)
        
        if not account_pairs:
            raise UserError(_("No debit-credit pairs were defined on row 6 (columns N-Y)."))
        
        missed_rows = []
        row_entries = []
        
        for row_idx in range(DATA_START_ROW_INDEX, len(df)):
            try:
                row_data = df.iloc[row_idx]
                personal_number_raw = self._safe_string(row_data.iloc[PERSONAL_NUMBER_COL_INDEX]) if len(row_data) > PERSONAL_NUMBER_COL_INDEX else ''
                personal_number = self._normalize_personal_number(personal_number_raw)
                excel_row = row_idx + 1
                
                if not personal_number:
                    missed_rows.append({
                        'excel_row': excel_row,
                        'personal_number': personal_number_raw or '',
                        'error': 'Missing personal number',
                    })
                    continue
                
                partner = self._get_partner_from_personal_number(personal_number)
                if not partner:
                    missed_rows.append({
                        'excel_row': excel_row,
                        'personal_number': personal_number,
                        'error': 'Employee not found (identification_id)',
                    })
                    continue
                
                row_entries.append({
                    'row_idx': row_idx,
                    'row_data': row_data,
                    'personal_number_raw': personal_number_raw,
                    'personal_number': personal_number,
                    'partner': partner,
                })
            except Exception as e:
                _logger.error("Error processing row %s: %s" % (row_idx, str(e)))
                missed_rows.append({
                    'excel_row': row_idx + 1,
                    'personal_number': '',
                    'error': str(e),
                })
        
        if missed_rows:
            return self._generate_and_download_missed_excel_payment(missed_rows)
        
        total_processed = 0
        row_number = 1
        for entry in row_entries:
            row_data = entry['row_data']
            personal_number_raw = entry['personal_number_raw']
            partner = entry['partner']
            for col_idx in AMOUNT_COLUMNS:
                if col_idx not in account_pairs or len(row_data) <= col_idx:
                    continue
                amount = self._safe_float(row_data.iloc[col_idx])
                if amount <= 0:
                    continue
                debit_code, credit_code = account_pairs[col_idx]
                column_letter = self._column_letter(col_idx)
                line_vals = {
                    'import_id': self.id,
                    'partner_name': personal_number_raw,
                    'partner_id': partner.id,
                    'net_amount': amount,
                    'debit': debit_code,
                    'credit': credit_code,
                    'credit_partner': False,
                    'project_name': False,
                    'column_description': column_labels.get(col_idx),
                    'row_number': row_number,
                    'excel_row': entry['row_idx'] + 1,
                }
                self.env['salary.payment.import.line'].create(line_vals)
                total_processed += 1
                row_number += 1
                _logger.info(f"Created line for partner {partner.name} column {column_letter} amount {amount}")
        
        if total_processed == 0:
            raise UserError(_("No payment lines were created. Please check Excel data (columns C and N-Y)."))
        
        self.state = 'imported'
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Success'),
                'message': _('Imported %d lines successfully!') % total_processed,
                'type': 'success',
            }
        }
    
    def _generate_and_download_missed_excel_payment(self, missed_rows):
        """Generate Excel file with missed rows and return download action."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise UserError(_("Generating missed-rows Excel requires openpyxl. Please install it."))
        stream = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Missed Rows'
        headers = ['Excel Row', 'Personal Number', 'Error']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for mr in missed_rows:
            ws.append([
                mr.get('excel_row'),
                mr.get('personal_number', ''),
                mr.get('error', ''),
            ])
        wb.save(stream)
        excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')
        filename = 'salary_payment_import_missed_rows_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': excel_data,
            'res_model': 'salary.payment.import',
            'res_id': self.id,
            'type': 'binary',
        })
        message = _("No records were created. %s row(s) could not be processed (employee not found by identification_id).") % len(missed_rows)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import failed'),
                'message': message,
                'type': 'danger',
                'sticky': False,
                'next': {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                }
            }
        }
    
    def _safe_string(self, val):
        """Convert value to string safely"""
        try:
            if pd.isna(val) or val is None:
                return ''
            if str(val).lower() == 'nan':
                return ''
            return str(val).strip()
        except:
            return ''
    
    def _safe_float(self, val):
        """Convert value to float safely"""
        try:
            if pd.isna(val) or val == '' or val is None:
                return 0.0
            val_str = str(val).strip()
            if not val_str:
                return 0.0
            # Remove non-numeric characters but keep digits, dots, and minus signs
            cleaned = ''.join(c for c in val_str if c.isdigit() or c in '.-')
            if not cleaned or cleaned == '.' or cleaned == '-':
                return 0.0
            return float(cleaned)
        except:
            return 0.0
    
    def _normalize_personal_number(self, value):
        value = self._safe_string(value)
        if not value:
            return ''
        value = value.split('.')[0]
        digits = ''.join(filter(str.isdigit, value))
        return digits.zfill(11) if digits else ''
    
    def _get_partner_from_personal_number(self, personal_number):
        if not personal_number:
            return False
        employee = self.env['hr.employee'].search([('identification_id', '=', personal_number)], limit=1)
        return employee.work_contact_id if employee and employee.work_contact_id else False
    
    def _column_letter(self, index):
        try:
            letter = ''
            while index >= 0:
                letter = chr(index % 26 + ord('A')) + letter
                index = index // 26 - 1
            return letter or ''
        except Exception:
            return ''
    
    def _account_exists(self, code):
        if not code:
            return False
        return bool(self.env['account.account'].search([('code', '=', code)], limit=1))
    
    def _find_partner_by_id(self, partner_id_str):
        """Find partner by VAT/ID number"""
        if not partner_id_str:
            return False
        
        # Clean and format ID number
        clean_id = partner_id_str.split('.')[0].zfill(11)
        
        # Search by VAT
        partner = self.env['res.partner'].search([('vat', '=', clean_id)], limit=1)
        if partner:
            return partner
        
        # Fallback search with original string
        partner = self.env['res.partner'].search([('vat', '=', partner_id_str)], limit=1)
        return partner
    
    def _find_partner_by_name(self, partner_name):
        """Find partner by name with multiple search strategies"""
        if not partner_name:
            return False
        
        clean_name = partner_name.strip()
        
        # Try different search strategies
        search_strategies = [
            [('name', '=', clean_name)],           # Exact match
            [('name', '=ilike', clean_name)],      # Case insensitive exact
            [('name', 'ilike', "%%%s%%" % clean_name)], # Contains
        ]
        
        for domain in search_strategies:
            partners = self.env['res.partner'].search(domain, limit=1)
            if partners:
                return partners
        
        # Create new partner if not found
        try:
            new_partner = self.env['res.partner'].create({
                'name': clean_name,
                'is_company': False,
                'supplier_rank': 1,
                'customer_rank': 1
            })
            return new_partner
        except:
            return False
    
    def action_generate_journal_entries(self):
        self.ensure_one()
        
        if self.state == 'posted':
            raise UserError(_("Journal entries already generated!"))
        
        if not self.line_ids:
            raise UserError(_("No lines to process!"))
        
        # Get general journal
        journal = self.env['account.journal'].search([('name', '=', 'Salaries')], limit=1)
        if not journal:
            raise UserError(_("General journal not found!"))
        
        journal_entries = self.env['account.move']
        counter = 1
        successful_entries = 0
        
        # Process each line individually
        for line in self.line_ids:
            try:
                # Must have partner for debit
                if not line.partner_id:
                    counter += 1
                    continue
                
                # Must have amount
                if line.net_amount <= 0:
                    counter += 1
                    continue
                
                amount = line.net_amount
                
                # GET DEBIT ACCOUNT FROM COLUMN C
                debit_account = None
                if line.debit and line.debit.strip():
                    debit_account = self.env['account.account'].search([('code', '=', line.debit.strip())], limit=1)
                    if not debit_account:
                        # Create account if doesn't exist
                        try:
                            debit_account = self.env['account.account'].create({
                                'name': 'Account %s' % line.debit.strip(),
                                'code': line.debit.strip(),
                                'account_type': 'liability_current',
                            })
                        except Exception as e:
                            _logger.error("Cannot create debit account %s: %s" % (line.debit.strip(), str(e)))
                            counter += 1
                            continue
                
                # GET CREDIT ACCOUNT FROM COLUMN D OR E
                credit_account = None
                credit_partner_for_entry = None
                
                # Priority 1: If Column E has partner, use their payable account
                if line.credit_partner_id:
                    credit_account = line.credit_partner_id.property_account_payable_id
                    credit_partner_for_entry = line.credit_partner_id
                    _logger.info("Using credit partner %s with account %s" % (line.credit_partner_id.name, credit_account.code if credit_account else 'None'))
                
                # Priority 2: If Column D has account code, use that account
                elif line.credit and line.credit.strip():
                    credit_account = self.env['account.account'].search([('code', '=', line.credit.strip())], limit=1)
                    if not credit_account:
                        # Create account if doesn't exist
                        try:
                            credit_account = self.env['account.account'].create({
                                'name': 'Account %s' % line.credit.strip(),
                                'code': line.credit.strip(),
                                'account_type': 'liability_current',
                            })
                        except Exception as e:
                            _logger.error("Cannot create credit account %s: %s" % (line.credit.strip(), str(e)))
                            counter += 1
                            continue
                    credit_partner_for_entry = None
                    _logger.info("Using credit account %s without partner" % credit_account.code)
                
                # Priority 3: Fallback to debit partner's payable account
                else:
                    credit_account = line.partner_id.property_account_payable_id
                    credit_partner_for_entry = line.partner_id
                    _logger.info("Using debit partner %s payable account as fallback" % line.partner_id.name)
                
                # Must have both accounts
                if not debit_account or not credit_account:
                    _logger.warning("Missing accounts - Debit: %s, Credit: %s" % (debit_account, credit_account))
                    counter += 1
                    continue
                
                # Prepare analytic distribution for project (ONLY for debit line)
                analytic_distribution = {}
                if line.analytic_account_id:
                    analytic_distribution[str(line.analytic_account_id.id)] = 100.0
                    _logger.info(f"Analytic distribution created: {analytic_distribution}")
                else:
                    _logger.info("No analytic account found for line")
                
                # Ensure analytic distribution is properly formatted for both lines
                if analytic_distribution:
                    analytic_distribution_debit = analytic_distribution.copy()
                    analytic_distribution_credit = analytic_distribution.copy()
                else:
                    analytic_distribution_debit = False
                    analytic_distribution_credit = False
                
                # Check if debit partner is a company
                is_company_partner = line.partner_id.is_company if line.partner_id else False
                _logger.info(f"Partner: {line.partner_id.name}, is_company: {is_company_partner}")
                _logger.info(f"Project name: {line.project_name}, Analytic account: {line.analytic_account_id.name if line.analytic_account_id else 'None'}")
                
                # Create journal entry with project support
                move_vals = {
                    'journal_id': journal.id,
                    'date': self.date,
                    'ref': _("Payment %d - %s - %.2f GEL%s") % (
                        counter, 
                        line.partner_id.name, 
                        amount,
                        " - Project: %s" % line.project_name if line.project_name else ""
                    ),
                    'line_ids': [
                        # Debit line (Column C account + Column A partner + Project)
                        (0, 0, {
                            'account_id': debit_account.id,
                            'partner_id': line.partner_id.id,
                            'analytic_distribution': analytic_distribution_debit,  # Analytic on debit
                            'debit': amount,
                            'credit': 0.0,
                            'name': _('Payment Debit - %s%s') % (
                                line.partner_id.name,
                                " - Project: %s" % line.project_name if line.project_name else ""
                            )
                        }),
                        # Credit line (Column D account or Column E partner + Project)
                        # For company partners, credit line should not have partner
                        (0, 0, {
                            'account_id': credit_account.id,
                            'partner_id': None if is_company_partner else (credit_partner_for_entry.id if credit_partner_for_entry else None),
                            'analytic_distribution': False,  # NO analytic on credit line
                            'debit': 0.0,
                            'credit': amount,
                            'name': _('Payment Credit - %s%s') % (
                                credit_partner_for_entry.name if credit_partner_for_entry and not is_company_partner else f'Account {credit_account.code}',
                                " - Project: %s" % line.project_name if line.project_name else ""
                            )
                        })
                    ]
                }
                
                # Create the journal entry
                move = self.env['account.move'].create(move_vals)
                journal_entries += move
                successful_entries += 1
                
                _logger.info("Created journal entry %s: Debit %s / Credit %s for %s - Project: %s - Company Partner: %s - Analytic: %s" % (counter, debit_account.code, credit_account.code, amount, line.project_name or 'N/A', is_company_partner, analytic_distribution))
                
            except Exception as e:
                _logger.error("Error creating journal entry for line %s: %s" % (counter, str(e)))
                pass
            
            counter += 1
        
        # Update state and link journal entries
        self.write({
            'journal_entry_ids': [(6, 0, journal_entries.ids)],
            'state': 'posted'
        })
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Success'),
                'message': _('%d journal entries created successfully!') % successful_entries,
                'type': 'success',
            }
        }
    
    def action_view_journal_entries(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Journal Entries'),
            'res_model': 'account.move',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.journal_entry_ids.ids)]
        }
    
    def action_create_journal_entries_3139(self):
        """
        Create a single journal entry per record for all lines.
        All lines are collected and added as account.move.line records to one account.move.
        Can handle multiple records (for server actions from list view).
        Returns action to view the created journal entries.
        """
        # Get general journal
        journal = self.env['account.journal'].search([('name', '=', 'Salaries')], limit=1)
        if not journal:
            raise UserError(_("General journal not found!"))
        
        all_journal_entries = self.env['account.move']
        total_successful_lines = 0
        processed_records = 0
        
        # Process each record in the recordset
        for record in self:
            if record.state == 'posted':
                _logger.warning("Skipping record %s - already posted" % record.name)
                continue
            
            if not record.line_ids:
                _logger.info("No lines found for record %s" % record.name)
                continue
            
            # Collect all move lines for this record
            all_move_lines = []
            successful_lines = 0
            
            # Process each line
            for line in record.line_ids:
                try:
                    # Must have partner for debit
                    if not line.partner_id:
                        continue
                    
                    # Must have amount
                    if line.net_amount <= 0:
                        continue
                    
                    amount = line.net_amount
                    
                    # Get debit account from line's debit field
                    debit_account = None
                    if line.debit and line.debit.strip():
                        debit_account = self.env['account.account'].search([('code', '=', line.debit.strip())], limit=1)
                        if not debit_account:
                            # Create account if doesn't exist
                            try:
                                debit_account = self.env['account.account'].create({
                                    'name': 'Account %s' % line.debit.strip(),
                                    'code': line.debit.strip(),
                                    'account_type': 'liability_current',
                                })
                            except Exception as e:
                                _logger.error("Cannot create debit account %s: %s" % (line.debit.strip(), str(e)))
                                continue
                    
                    if not debit_account:
                        _logger.warning("Missing debit account for line %s" % line.id)
                        continue
                    
                    # Get credit account from Column D or E
                    credit_account = None
                    credit_partner_for_entry = None
                    
                    # Priority 1: If Column E has partner, use their payable account
                    if line.credit_partner_id:
                        credit_account = line.credit_partner_id.property_account_payable_id
                        credit_partner_for_entry = line.credit_partner_id
                        _logger.info("Using credit partner %s with account %s" % (line.credit_partner_id.name, credit_account.code if credit_account else 'None'))
                    
                    # Priority 2: If Column D has account code, use that account
                    elif line.credit and line.credit.strip():
                        credit_account = self.env['account.account'].search([('code', '=', line.credit.strip())], limit=1)
                        if not credit_account:
                            # Create account if doesn't exist
                            try:
                                credit_account = self.env['account.account'].create({
                                    'name': 'Account %s' % line.credit.strip(),
                                    'code': line.credit.strip(),
                                    'account_type': 'liability_current',
                                })
                            except Exception as e:
                                _logger.error("Cannot create credit account %s: %s" % (line.credit.strip(), str(e)))
                                continue
                        credit_partner_for_entry = None
                        _logger.info("Using credit account %s without partner" % credit_account.code)
                    
                    # Priority 3: Fallback to debit partner's payable account
                    else:
                        credit_account = line.partner_id.property_account_payable_id
                        credit_partner_for_entry = line.partner_id
                        _logger.info("Using debit partner %s payable account as fallback" % line.partner_id.name)
                    
                    # Must have both accounts
                    if not debit_account or not credit_account:
                        _logger.warning("Missing accounts - Debit: %s, Credit: %s" % (debit_account, credit_account))
                        continue
                    
                    # Prepare analytic distribution for project (ONLY for debit line)
                    analytic_distribution = {}
                    if line.analytic_account_id:
                        analytic_distribution[str(line.analytic_account_id.id)] = 100.0
                        _logger.info(f"Analytic distribution created: {analytic_distribution}")
                    
                    # Ensure analytic distribution is properly formatted
                    if analytic_distribution:
                        analytic_distribution_debit = analytic_distribution.copy()
                        analytic_distribution_credit = False
                    else:
                        analytic_distribution_debit = False
                        analytic_distribution_credit = False
                    
                    # Set partner_id based on account code: only if account code is "3139" or "3323"
                    # For debit line
                    if debit_account.code in ['3139', '3323']:
                        debit_partner_id = line.partner_id.id
                    else:
                        debit_partner_id = False
                    
                    # For credit line
                    if credit_account.code in ['3139', '3323']:
                        # If credit account is 3139 or 3323, use credit_partner_for_entry if available, otherwise debit partner
                        if credit_partner_for_entry:
                            credit_partner_id = credit_partner_for_entry.id
                        else:
                            credit_partner_id = line.partner_id.id
                    else:
                        credit_partner_id = False
                    
                    # Prepare move lines
                    project_suffix = " - Project: %s" % line.project_name if line.project_name else ""
                    
                    # Debit line
                    all_move_lines.append((0, 0, {
                        'account_id': debit_account.id,
                        'partner_id': debit_partner_id,
                        'analytic_distribution': analytic_distribution_debit,
                        'debit': amount,
                        'credit': 0.0,
                        'name': _('Payment Debit - %s%s') % (
                            line.partner_id.name,
                            project_suffix
                        )
                    }))
                    
                    # Credit line
                    credit_name = credit_partner_for_entry.name if credit_partner_for_entry else f'Account {credit_account.code}'
                    all_move_lines.append((0, 0, {
                        'account_id': credit_account.id,
                        'partner_id': credit_partner_id,
                        'analytic_distribution': False,
                        'debit': 0.0,
                        'credit': amount,
                        'name': _('Payment Credit - %s%s') % (
                            credit_name,
                            project_suffix
                        )
                    }))
                    
                    successful_lines += 2  # Count both debit and credit lines
                    _logger.info("Added move line: Debit %s / Credit %s for %s - Project: %s" % (
                        debit_account.code, credit_account.code, amount, line.project_name or 'N/A'
                    ))
                    
                except Exception as e:
                    _logger.error("Error preparing move line for line %s: %s" % (line.id, str(e)))
                    continue
            
            # Create a single journal entry with all collected lines
            if all_move_lines:
                try:
                    move = self.env['account.move'].create({
                        'journal_id': journal.id,
                        'date': record.date,
                        'ref': _("Payment - %s") % record.name,
                        'line_ids': all_move_lines
                    })
                    
                    # Link journal entry to record
                    existing_entry_ids = record.journal_entry_ids.ids
                    new_entry_ids = [move.id]
                    all_entry_ids = list(set(existing_entry_ids + new_entry_ids))
                    
                    record.write({
                        'journal_entry_ids': [(6, 0, all_entry_ids)],
                        'state': 'posted'
                    })
                    
                    all_journal_entries += move
                    total_successful_lines += successful_lines
                    processed_records += 1
                    _logger.info("Processed record %s: Created 1 journal entry with %d move lines" % (record.name, successful_lines))
                    
                except Exception as e:
                    _logger.error("Error creating journal entry for record %s: %s" % (record.name, str(e)))
                    raise UserError(_("Error creating journal entry for record %s: %s") % (record.name, str(e)))
            else:
                _logger.warning("No move lines to create for record %s" % record.name)
        
        if total_successful_lines == 0:
            raise UserError(_("No journal entries were created. Please check your data."))
        
        # Return action to view all created journal entries
        return {
            'type': 'ir.actions.act_window',
            'name': _('Journal Entries'),
            'res_model': 'account.move',
            'view_mode': 'list,form',
            'domain': [('id', 'in', all_journal_entries.ids)],
            'context': {'create': False}
        }


class SalaryPaymentImportLine(models.Model):
    _name = 'salary.payment.import.line'
    _description = 'Salary Payment Import Line'

    import_id = fields.Many2one('salary.payment.import', required=True, ondelete='cascade')
    partner_name = fields.Char(string='A სვეტი - პირადი ნომერი')
    partner_id = fields.Many2one('res.partner', string='დებიტის პარტნიორი (A)')
    net_amount = fields.Float(string='თანხა (B)')
    debit = fields.Char(string='დებიტი (C)')
    credit = fields.Char(string='კრედიტი (D)')
    credit_partner = fields.Char(string='E სვეტი - კრედიტის პარტნიორის სახელი')
    credit_partner_id = fields.Many2one('res.partner', string='კრედიტის პარტნიორი (E)')
    project_name = fields.Char(string='F სვეტი - PROJECT')
    column_description = fields.Char(string='Column label (row 4)')
    analytic_account_id = fields.Many2one('account.analytic.account', string='პროექტი/Analytic Account')
    row_number = fields.Integer(string='Line Number')
    excel_row = fields.Integer(string='Excel Row')
    state = fields.Selection(related='import_id.state', store=True)
    company_id = fields.Many2one(related='import_id.company_id', store=True)