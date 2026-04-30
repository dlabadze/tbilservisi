from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import pandas as pd
import io
import logging
import traceback
import mimetypes

_logger = logging.getLogger(__name__)

class SalaryImport(models.Model):
    _name = 'salary.import'
    _description = 'Salary Import'
    _order = 'create_date desc'

    name = fields.Char(string='Reference', required=True, copy=False, default='New')
    date = fields.Date(string='Date', default=fields.Date.context_today, required=True)
    excel_file = fields.Binary(string='Excel File', required=True)
    excel_filename = fields.Char(string='File Name')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('imported', 'Imported'),
        ('posted', 'Posted')
    ], string='Status', default='draft', copy=False)
    line_ids = fields.One2many('salary.import.line', 'import_id', string='Salary Lines')
    journal_entry_count = fields.Integer(compute='_compute_journal_entry_count')
    journal_entry_ids = fields.Many2many('account.move', string='Journal Entries', copy=False)
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    department_id = fields.Many2one('hr.department', string='Department')
    
    def _compute_journal_entry_count(self):
        for record in self:
            record.journal_entry_count = len(record.journal_entry_ids)
    
    @api.model
    def create(self, vals):
        if vals.get('name', 'New') == 'New':
            vals['name'] = self.env['ir.sequence'].next_by_code('salary.import') or 'New'
        return super(SalaryImport, self).create(vals)
    
    def _validate_excel_file(self, excel_data, filename):
        """
        Validate Excel file before processing
        """
        # Check file size
        if len(excel_data) == 0:
            raise UserError("The uploaded file is empty.")
        
        # Check file type
        if filename:
            mime_type, _ = mimetypes.guess_type(filename)
            valid_types = [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                'application/vnd.ms-excel',  # .xls
            ]
            if mime_type not in valid_types:
                raise UserError(f"Invalid file type. Please upload an Excel file. Detected type: {mime_type}")
        
        return True
    
    def action_import_excel(self):
        for record in self:
            if not record.excel_file:
                raise UserError("Please upload an Excel file.")
            
            # Decode the file
            try:
                excel_data = base64.b64decode(record.excel_file)
            except Exception as e:
                _logger.error(f"Base64 Decoding Error: {str(e)}")
                raise UserError("Error decoding the uploaded file. Please check the file and try again.")
            
            # Validate file
            self._validate_excel_file(excel_data, record.excel_filename)
            
            # Prepare file for reading
            file_buffer = io.BytesIO(excel_data)
            
            # Read row 6 (index 5) for debit/credit information
            def _read_row6(read_engine):
                file_buffer.seek(0)
                df_row6 = pd.read_excel(
                    file_buffer,
                    engine=read_engine,
                    header=None,
                    nrows=6,  # Read first 6 rows
                    usecols="D:M",  # Columns D through M (inclusive)
                    dtype=str
                )
                # Return row 6 (index 5)
                if len(df_row6) > 5:
                    return df_row6.iloc[5]  # Row 6 (0-based index 5)
                return pd.Series()
            
            # Try reading with multiple approaches
            def _read_dataframe(read_engine):
                file_buffer.seek(0)
                return pd.read_excel(
                    file_buffer,
                    engine=read_engine,
                    header=None,
                    skiprows=6,  # First 6 rows are headers/info, data starts at row 7
                    usecols="B:M",  # Columns B through M (inclusive)
                    dtype=str
                )
            
            # Read row 6 for debit/credit values
            row6_data = None
            try:
                row6_data = _read_row6('openpyxl')
            except Exception as e:
                try:
                    row6_data = _read_row6('xlrd')
                except Exception as e2:
                    _logger.warning(f"Could not read row 6 for debit/credit: {e2}")
                    row6_data = pd.Series()
            
            # Extract debit/credit values from row 6
            def safe_string_from_row6(index):
                if row6_data is None or len(row6_data) <= index:
                    return ''
                val = row6_data.iloc[index] if hasattr(row6_data, 'iloc') else (row6_data[index] if index < len(row6_data) else '')
                if pd.isna(val) or val is None:
                    return ''
                return str(val).strip()
            
            # Map row 6 columns D-M (indices 0-9) to debit_credit fields
            debit_credit_values = {
                'base': safe_string_from_row6(0),          # Column D
                'vacation': safe_string_from_row6(1),      # Column E
                'bulletin': safe_string_from_row6(2),      # Column F
                'bonus': safe_string_from_row6(3),         # Column G
                'help': safe_string_from_row6(4),          # Column H
                'overtime': safe_string_from_row6(5),      # Column I
                'excessive_overtime_ded': safe_string_from_row6(6),  # Column J
                'pension': safe_string_from_row6(7),       # Column K
                'company': safe_string_from_row6(8),       # Column L
                'income': safe_string_from_row6(9),        # Column M
            }
            
            try:
                df = _read_dataframe('openpyxl')
            except Exception as first_error:
                try:
                    df = _read_dataframe('xlrd')
                except Exception as second_error:
                    _logger.error(f"First attempt error: {first_error}")
                    _logger.error(f"Second attempt error: {second_error}")
                    _logger.error(f"Traceback: {traceback.format_exc()}")
                    raise UserError(
                        "Could not read the Excel file. Possible reasons:\n"
                        "- File is corrupted\n"
                        "- File is in an unsupported format\n"
                        "- File requires password\n"
                        "- Incompatible Excel version"
                    )
            
            if df is None or df.empty:
                raise UserError("The Excel file does not contain any data after the header rows.")
            
            # Drop fully empty rows
            df = df.dropna(how='all')
            if df.empty:
                raise UserError("All rows are empty after removing headers.")
            
            # Clear existing lines
            record.line_ids.unlink()
            
            COLUMN_INDEX = {
                'full_name': 0,                     # Column B
                'personal_number': 1,               # Column C
                'salary': 2,                        # Column D
                'vacation_salary': 3,               # Column E
                'bulletin_salary': 4,               # Column F
                'bonus_salary': 5,                  # Column G
                'help_salary': 6,                   # Column H
                'overtime_salary': 7,               # Column I
                'excessive_overtime_deduction': 8,  # Column J
                'pension': 9,                       # Column K
                'company_tax': 10,                  # Column L
                'income_tax': 11,                   # Column M
            }
            
            def safe_float(val):
                try:
                    if pd.isna(val) or val is None:
                        return 0.0
                    if isinstance(val, str):
                        cleaned = ''.join(c for c in val if c.isdigit() or c in '.-')
                        val = cleaned or '0'
                    return float(val)
                except Exception:
                    return 0.0
            
            def safe_string(val):
                if pd.isna(val) or val is None:
                    return ''
                return str(val).strip()
            
            def normalize_personal_number(val):
                value = safe_string(val)
                if not value:
                    return ''
                value = value.split('.')[0]
                digits = ''.join(filter(str.isdigit, value))
                return digits.zfill(11) if digits else ''
            
            import_lines = []
            missed_rows = []
            
            for idx, row in df.iterrows():
                excel_row = int(idx) + 7  # data starts at row 7 (6 header rows)
                full_name = safe_string(row.iloc[COLUMN_INDEX['full_name']] if len(row) > COLUMN_INDEX['full_name'] else '')
                personal_number = normalize_personal_number(
                    row.iloc[COLUMN_INDEX['personal_number']] if len(row) > COLUMN_INDEX['personal_number'] else ''
                )
                if not personal_number:
                    missed_rows.append({
                        'excel_row': excel_row,
                        'personal_number': '',
                        'full_name': full_name,
                        'error': 'Missing personal number',
                    })
                    continue
                
                employee = self.env['hr.employee'].search([('identification_id', '=', personal_number)], limit=1)
                partner = employee.work_contact_id if employee else False
                if not partner:
                    missed_rows.append({
                        'excel_row': excel_row,
                        'personal_number': personal_number,
                        'full_name': full_name,
                        'error': 'Employee not found (identification_id)',
                    })
                    continue
                salary_value = safe_float(row.iloc[COLUMN_INDEX['salary']] if len(row) > COLUMN_INDEX['salary'] else 0.0)
                vacation_salary_value = safe_float(row.iloc[COLUMN_INDEX['vacation_salary']] if len(row) > COLUMN_INDEX['vacation_salary'] else 0.0)
                bulletin_salary_value = safe_float(row.iloc[COLUMN_INDEX['bulletin_salary']] if len(row) > COLUMN_INDEX['bulletin_salary'] else 0.0)
                bonus_salary_value = safe_float(row.iloc[COLUMN_INDEX['bonus_salary']] if len(row) > COLUMN_INDEX['bonus_salary'] else 0.0)
                help_salary_value = safe_float(row.iloc[COLUMN_INDEX['help_salary']] if len(row) > COLUMN_INDEX['help_salary'] else 0.0)
                overtime_salary_value = safe_float(row.iloc[COLUMN_INDEX['overtime_salary']] if len(row) > COLUMN_INDEX['overtime_salary'] else 0.0)
                excessive_overtime_deduction_value = safe_float(
                    row.iloc[COLUMN_INDEX['excessive_overtime_deduction']]
                    if len(row) > COLUMN_INDEX['excessive_overtime_deduction'] else 0.0
                )
                pension_value = safe_float(row.iloc[COLUMN_INDEX['pension']] if len(row) > COLUMN_INDEX['pension'] else 0.0)
                company_tax_value = safe_float(row.iloc[COLUMN_INDEX['company_tax']] if len(row) > COLUMN_INDEX['company_tax'] else 0.0)
                income_tax_value = safe_float(row.iloc[COLUMN_INDEX['income_tax']] if len(row) > COLUMN_INDEX['income_tax'] else 0.0)
                
                total_salary_value_1 = salary_value + vacation_salary_value + bulletin_salary_value + bonus_salary_value + help_salary_value\
                    + overtime_salary_value
                total_salary_value = total_salary_value_1 - excessive_overtime_deduction_value

                line_vals = {
                    'import_id': record.id,
                    'employee_id': partner.id if partner else False,
                    'partner_id': partner.id if partner else False,
                    'partner_vat': personal_number,
                    'full_name': full_name,
                    'base_salary': salary_value,
                    'vacation_salary': vacation_salary_value,
                    'bulletin_salary': bulletin_salary_value,
                    'bonus_salary': bonus_salary_value,
                    'help_salary': help_salary_value,
                    'overtime_salary': overtime_salary_value,
                    'excessive_overtime_deduction': excessive_overtime_deduction_value,
                    'total_salary': total_salary_value,
                    'company_tax': company_tax_value,
                    'pension': pension_value,
                    'income_tax': income_tax_value,
                    'net_amount': 0.0,
                    # Add debit/credit fields from row 6
                    'base_debit_credit': debit_credit_values['base'],
                    'vacation_debit_credit': debit_credit_values['vacation'],
                    'bulletin_debit_credit': debit_credit_values['bulletin'],
                    'bonus_debit_credit': debit_credit_values['bonus'],
                    'help_debit_credit': debit_credit_values['help'],
                    'overtime_debit_credit': debit_credit_values['overtime'],
                    'excessive_overtime_ded_debit_credit': debit_credit_values['excessive_overtime_ded'],
                    'pension_debit_credit': debit_credit_values['pension'],
                    'company_debit_credit': debit_credit_values['company'],
                    'income_debit_credit': debit_credit_values['income'],
                }
                
                import_lines.append(line_vals)
            
            if missed_rows:
                return self._generate_and_download_missed_excel_salary(record, missed_rows)
            
            # Batch create lines
            if import_lines:
                self.env['salary.import.line'].create(import_lines)
            
            # Update state
            record.state = 'imported'
            
            # Log import details
            _logger.info(f"Imported {len(import_lines)} salary lines")
        
        return True

    def _generate_and_download_missed_excel_salary(self, record, missed_rows):
        """Generate Excel file with missed rows and return download action."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise UserError("Generating missed-rows Excel requires openpyxl. Please install it.")
        stream = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Missed Rows'
        headers = ['Excel Row', 'Personal Number', 'Full Name', 'Error']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for mr in missed_rows:
            ws.append([
                mr.get('excel_row'),
                mr.get('personal_number', ''),
                mr.get('full_name', ''),
                mr.get('error', ''),
            ])
        wb.save(stream)
        excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')
        filename = 'salary_import_missed_rows_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': excel_data,
            'res_model': 'salary.import',
            'res_id': record.id,
            'type': 'binary',
        })
        message = "No records were created. %s row(s) could not be processed (employee not found by identification_id)." % len(missed_rows)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import failed',
                'message': message,
                'type': 'danger',
                'sticky': False,
                'next': {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                }
            }
        }

    def action_generate_journal_entries(self):
        for record in self:
            # Function to find account with fallback methods
            def find_account(code, name=None):
                # Try finding by code without company filter
                account = self.env['account.account'].search([
                    ('code', '=', code)
                ], limit=1)
                
                # If still not found and name is provided, create the account
                if not account and name:
                    try:
                        account = self.env['account.account'].create({
                            'name': name,
                            'code': code,
                            'account_type': 'liability_current',  # Adjust account type as needed
                        })
                    except Exception as e:
                        _logger.error(f"Error creating account {code}: {e}")
                        raise UserError(f"Could not create account {code}: {e}")
                
                return account
            
            # Find accounts 
            accounts = {
                '7410': find_account('7410', 'Salary Expense Account'),
                '3130': find_account('3130', 'Salary Payable Account'),
                '3370': find_account('3370', 'General Liability Account'),
                '7415': find_account('7415', 'Company Tax Expense Account'),
                '3320': find_account('3320', 'Income Tax Liability Account')
            }
            
            # Get general journal
            journal = self.env['account.journal'].search([('name', '=', 'Salaries')], limit=1)
            
            # Verify accounts and journal exist
            missing_accounts = [code for code, account in accounts.items() if not account]
            if missing_accounts or not journal:
                missing = missing_accounts + (['General Journal'] if not journal else [])
                raise UserError(f"Required accounts or journal not found: {', '.join(missing)}")
            
            # Prepare journal entries
            journal_entries = self.env['account.move']
            
            # Process each line
            for line in record.line_ids:
                if not line.partner_id:
                    continue
                
                try:
                    # Create a single journal entry with all line items
                    move_lines = []
                    
                    # 1. Main salary entry
                    move_lines.extend([
                        # Salary expense debit (7410)
                        (0, 0, {
                            'account_id': accounts['7410'].id,
                            'partner_id': line.partner_id.id,
                            'debit': line.total_salary,
                            'credit': 0,
                            'name': f"{line.partner_id.name} - Salary Expense"
                        }),
                        # Salary credit (3130)
                        (0, 0, {
                            'account_id': accounts['3130'].id,
                            'partner_id': line.partner_id.id,
                            'debit': 0,
                            'credit': line.total_salary,
                            'name': f"{line.partner_id.name} - Salary Payable"
                        })
                    ])
                    
                    # 2. Pension entries if applicable
                    if line.pension > 0:
                        move_lines.extend([
                            # Pension debit (3130)
                            (0, 0, {
                                'account_id': accounts['3130'].id,
                                'partner_id': line.partner_id.id,
                                'debit': line.pension,
                                'credit': 0,
                                'name': f"{line.partner_id.name} - Pension"
                            }),
                            # Pension credit (3370)
                            (0, 0, {
                                'account_id': accounts['3370'].id,
                                'partner_id': line.partner_id.id,
                                'debit': 0,
                                'credit': line.pension,
                                'name': f"{line.partner_id.name} - Pension Liability"
                            })
                        ])
                    
                    # 3. Company tax entries if applicable
                    if line.company_tax > 0:
                        move_lines.extend([
                            # Company tax debit (7415)
                            (0, 0, {
                                'account_id': accounts['7415'].id,
                                'partner_id': line.partner_id.id,
                                'debit': line.company_tax,
                                'credit': 0,
                                'name': f"{line.partner_id.name} - Company Tax"
                            }),
                            # Company tax credit (3370)
                            (0, 0, {
                                'account_id': accounts['3370'].id,
                                'partner_id': line.partner_id.id,
                                'debit': 0,
                                'credit': line.company_tax,
                                'name': f"{line.partner_id.name} - Company Tax Liability"
                            })
                        ])
                    
                    # 4. Income tax entries if applicable
                    if line.income_tax > 0:
                        move_lines.extend([
                            # Income tax debit (3130)
                            (0, 0, {
                                'account_id': accounts['3130'].id,
                                'partner_id': line.partner_id.id,
                                'debit': line.income_tax,
                                'credit': 0,
                                'name': f"{line.partner_id.name} - Income Tax"
                            }),
                            # Income tax credit (3320)
                            (0, 0, {
                                'account_id': accounts['3320'].id,
                                'partner_id': line.partner_id.id,
                                'debit': 0,
                                'credit': line.income_tax,
                                'name': f"{line.partner_id.name} - Income Tax Liability"
                            })
                        ])
                    
                    # Create the journal entry
                    move = self.env['account.move'].create({
                        'ref': f"Salary for {line.partner_id.name}",
                        'date': record.date,
                        'journal_id': journal.id,
                        'line_ids': move_lines
                    })
                    
                    journal_entries += move
                    
                except Exception as e:
                    _logger.error(f"Error creating journal entry for {line.partner_id.name}: {e}")
                    raise UserError(f"Error creating journal entry for {line.partner_id.name}: {e}")
            
            # Link journal entries to salary import record
            record.journal_entry_ids = journal_entries
            record.state = 'posted'
            
            return True

    def action_view_journal_entries(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Journal Entries',
            'res_model': 'account.move',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.journal_entry_ids.ids)]
        }
    
    def action_generate_journal_entries_by_field(self):
        """
        Generate a single journal entry per record with all salary field lines.
        Each field with value > 0 will create account.move.line records using its corresponding debit_credit field.
        Can handle multiple records (for server actions from list view).
        """
        # Get general journal (once for all records)
        journal = self.env['account.journal'].search([('name', '=', 'Salaries')], limit=1)
        if not journal:
            raise UserError("General journal not found!")
        
        def find_account(code):
            """Find account by code, create if doesn't exist"""
            if not code or not code.strip():
                return False
            account = self.env['account.account'].search([('code', '=', code.strip())], limit=1)
            if not account:
                try:
                    account = self.env['account.account'].create({
                        'name': f'Account {code.strip()}',
                        'code': code.strip(),
                        'account_type': 'liability_current',
                    })
                except Exception as e:
                    _logger.error(f"Error creating account {code}: {e}")
                    return False
            return account
        
        def parse_debit_credit(debit_credit_str):
            """Parse debit-credit string (format: 'debit-credit' or 'debit/credit')"""
            if not debit_credit_str or not debit_credit_str.strip():
                return None, None
            
            # Try different separators
            for sep in ['-', '/', '–']:
                if sep in debit_credit_str:
                    parts = [p.strip() for p in debit_credit_str.split(sep)]
                    if len(parts) >= 2:
                        return parts[0], parts[1]
            
            # If no separator, check if it's just a single account code
            # Check if it contains specific account codes
            if '3133.28' in str(debit_credit_str):
                return '3133.28', None
            if '3139' in str(debit_credit_str):
                return '3139', None
            
            return None, None
        
        # Field mapping: salary_field -> debit_credit_field
        field_mapping = {
            'base_salary': 'base_debit_credit',
            'vacation_salary': 'vacation_debit_credit',
            'bulletin_salary': 'bulletin_debit_credit',
            'bonus_salary': 'bonus_debit_credit',
            'help_salary': 'help_debit_credit',
            'overtime_salary': 'overtime_debit_credit',
            'excessive_overtime_deduction': 'excessive_overtime_ded_debit_credit',
            'pension': 'pension_debit_credit',
            'company_tax': 'company_debit_credit',
            'income_tax': 'income_debit_credit',
        }
        
        all_journal_entries = self.env['account.move']
        total_successful_lines = 0
        processed_records = 0
        skipped_records = 0
        
        # Process each record
        for record in self:
            # Skip already posted records
            if record.state == 'posted':
                _logger.info(f"Skipping record {record.name} - already posted")
                skipped_records += 1
                continue
            
            if not record.line_ids:
                _logger.warning(f"Skipping record {record.name} - no lines to process")
                skipped_records += 1
                continue
            
            # Collect all move lines for this record
            all_move_lines = []
            successful_lines = 0
            
            # Process each line
            for line in record.line_ids:
                if not line.partner_id:
                    continue
                
                # Process each salary field
                for salary_field, debit_credit_field in field_mapping.items():
                    salary_value = getattr(line, salary_field, 0.0)
                    
                    # Skip if value is 0 or less
                    if salary_value <= 0:
                        continue
                    
                    debit_credit_str = getattr(line, debit_credit_field, '')
                    
                    # Skip if no debit/credit information
                    if not debit_credit_str or not debit_credit_str.strip():
                        _logger.warning(f"Skipping {salary_field} for {line.partner_id.name}: no debit/credit info")
                        continue
                    
                    # Parse debit and credit accounts
                    debit_code, credit_code = parse_debit_credit(debit_credit_str)
                    
                    # If debit_credit contains specific account codes, use them as debit
                    if '3133.28' in str(debit_credit_str) and not debit_code:
                        debit_code = '3133.28'
                    if '3139' in str(debit_credit_str) and not debit_code:
                        debit_code = '3139'
                    
                    if not debit_code:
                        _logger.warning(f"Skipping {salary_field} for {line.partner_id.name}: could not parse debit account")
                        continue
                    
                    # Find or create accounts
                    debit_account = find_account(debit_code)
                    if not debit_account:
                        _logger.warning(f"Skipping {salary_field} for {line.partner_id.name}: debit account {debit_code} not found")
                        continue
                    
                    # For credit account, use credit_code if provided, otherwise use partner's payable account
                    if credit_code:
                        credit_account = find_account(credit_code)
                        if not credit_account:
                            _logger.warning(f"Skipping {salary_field} for {line.partner_id.name}: credit account {credit_code} not found")
                            continue
                    else:
                        # Use partner's payable account
                        credit_account = line.partner_id.property_account_payable_id
                        if not credit_account:
                            _logger.warning(f"Skipping {salary_field} for {line.partner_id.name}: partner has no payable account")
                            continue
                    
                    # Set partner_id based on account code: only if account code is "3139", "3323", or "3133.28"
                    # For debit line
                    if debit_account.code in ['3139', '3323', '3133.28']:
                        debit_partner_id = line.partner_id.id
                    else:
                        debit_partner_id = False
                    
                    # For credit line
                    if credit_account.code in ['3139', '3323', '3133.28']:
                        credit_partner_id = line.partner_id.id
                    else:
                        credit_partner_id = False
                    
                    # Prepare name for journal entry lines
                    partner_name = line.partner_id.name if (debit_partner_id or credit_partner_id) else ''
                    name_suffix = f" - {partner_name}" if partner_name else ""
                    
                    try:
                        # Prepare move lines
                        field_name = salary_field.replace('_', ' ').title()
                        
                        # Debit line
                        all_move_lines.append((0, 0, {
                            'account_id': debit_account.id,
                            'partner_id': debit_partner_id,
                            'debit': salary_value,
                            'credit': 0.0,
                            'name': f"{field_name} (Debit){name_suffix}"
                        }))
                        
                        # Credit line
                        all_move_lines.append((0, 0, {
                            'account_id': credit_account.id,
                            'partner_id': credit_partner_id,
                            'debit': 0.0,
                            'credit': salary_value,
                            'name': f"{field_name} (Credit){name_suffix}"
                        }))
                        
                        successful_lines += 2  # Count both debit and credit lines
                        _logger.info(f"Added move line for {line.partner_id.name} - {field_name}: {salary_value}")
                        
                    except Exception as e:
                        _logger.error(f"Error preparing move line for {line.partner_id.name} - {salary_field}: {e}")
                        continue
            
            # Create a single journal entry with all collected lines
            if all_move_lines:
                try:
                    move = self.env['account.move'].create({
                        'journal_id': journal.id,
                        'date': record.date,
                        'ref': f"Salary Import - {record.name}",
                        'line_ids': all_move_lines
                    })
                    
                    # Link journal entry to salary import record
                    existing_entry_ids = record.journal_entry_ids.ids
                    new_entry_ids = [move.id]
                    all_entry_ids = list(set(existing_entry_ids + new_entry_ids))
                    
                    record.write({
                        'journal_entry_ids': [(6, 0, all_entry_ids)],
                        'state': 'posted'
                    })
                    
                    all_journal_entries += move
                    total_successful_lines += successful_lines
                    processed_records += 1
                    _logger.info(f"Processed record {record.name}: Created 1 journal entry with {successful_lines} move lines")
                    
                except Exception as e:
                    _logger.error(f"Error creating journal entry for record {record.name}: {e}")
                    raise UserError(f"Error creating journal entry for record {record.name}: {e}")
            else:
                _logger.warning(f"No move lines to create for record {record.name}")
        
        # Return summary
        if total_successful_lines > 0:
            message = f'Created {processed_records} journal entry(ies) with {total_successful_lines} move lines'
            if skipped_records > 0:
                message += f' ({skipped_records} record(s) skipped - already posted or no lines)'
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Success',
                    'message': message,
                    'type': 'success',
                }
            }
        else:
            if skipped_records == len(self):
                raise UserError("All selected records are already posted or have no lines to process.")
            else:
                raise UserError("No journal entries were created. Please check your data.")


class SalaryImportLine(models.Model):
    _name = 'salary.import.line'
    _description = 'Salary Import Line'

    import_id = fields.Many2one('salary.import', required=True, ondelete='cascade')
    employee_id = fields.Many2one('res.partner', string='Employee')
    partner_id = fields.Many2one('res.partner', string='Employee')
    partner_vat = fields.Char(string='ID Number')
    full_name = fields.Char(string='Full Name')
    base_salary = fields.Float(string='Base Salary')
    total_salary = fields.Float(string='Total Salary')
    pension = fields.Float(string='Pension (2%)')
    company_tax = fields.Float(string='Company Tax')
    income_tax = fields.Float(string='Income Tax')
    net_amount = fields.Float(string='Net Amount')
    state = fields.Selection(related='import_id.state', store=True)
    company_id = fields.Many2one(related='import_id.company_id', store=True)
    vacation_salary = fields.Float(string='Vacation Salary')
    bulletin_salary = fields.Float(string='Bulletin Salary')
    bonus_salary = fields.Float(string='Bonus Salary')
    help_salary = fields.Float(string='Help Salary')
    overtime_salary = fields.Float(string='Overtime Salary')
    excessive_overtime_deduction = fields.Float(string='Excessive Overtime Deduction')
    # Debit/Credit fields from row 6
    base_debit_credit = fields.Char(string='Base Debit/Credit')
    vacation_debit_credit = fields.Char(string='Vacation Debit/Credit')
    bulletin_debit_credit = fields.Char(string='Bulletin Debit/Credit')
    bonus_debit_credit = fields.Char(string='Bonus Debit/Credit')
    help_debit_credit = fields.Char(string='Help Debit/Credit')
    overtime_debit_credit = fields.Char(string='Overtime Debit/Credit')
    excessive_overtime_ded_debit_credit = fields.Char(string='Excessive Overtime Ded Debit/Credit')
    pension_debit_credit = fields.Char(string='Pension Debit/Credit')
    company_debit_credit = fields.Char(string='Company Debit/Credit')
    income_debit_credit = fields.Char(string='Income Debit/Credit')