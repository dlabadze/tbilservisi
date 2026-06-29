# -*- coding: utf-8 -*-
from odoo import fields, models
from odoo.exceptions import UserError
import base64
from io import BytesIO


class PensionGadaxurvaWizard(models.TransientModel):
    _name = 'pension.gadaxurva.wizard'
    _description = 'საპენსიო გადახურვა'

    date = fields.Date(string='თარიღი', required=True, default=fields.Date.context_today)
    excel_file = fields.Binary(string='Excel ფაილი', required=True)
    excel_filename = fields.Char(string='ფაილის სახელი')

    def action_generate(self):
        self.ensure_one()
        data = base64.b64decode(self.excel_file)
        rows = self._read_excel(data)
        if not rows:
            raise UserError('ფაილში მონაცემები ვერ მოიძებნა.')

        account_debit = self.env['account.account'].search(
            [('code', '=', '3133.28')], limit=1
        )
        account_credit = self.env['account.account'].search(
            [('code', '=', '3133.27')], limit=1
        )
        if not account_debit:
            raise UserError('ანგარიში 3133.28 ვერ მოიძებნა.')
        if not account_credit:
            raise UserError('ანგარიში 3133.27 ვერ მოიძებნა.')

        journal = self.env['account.journal'].search(
            [('name', '=', 'Miscellaneous Operations')], limit=1
        )
        if not journal:
            raise UserError('საერთო ჟურნალი ვერ მოიძებნა.')

        line_vals = []
        missed_rows = []

        for row in rows:
            partner = self._find_partner(row['partner_name'], row['partner_vat'])
            if not partner:
                missed_rows.append(row)
                continue

            line_vals.append((0, 0, {
                'account_id': account_debit.id,
                'partner_id': partner.id,
                'debit': row['amount'],
                'credit': 0.0,
                'name': row['partner_name'],
            }))
            line_vals.append((0, 0, {
                'account_id': account_credit.id,
                'partner_id': False,
                'debit': 0.0,
                'credit': row['amount'],
                'name': row['partner_name'],
            }))

        move = False
        if line_vals:
            move = self.env['account.move'].create({
                'date': self.date,
                'journal_id': journal.id,
                'line_ids': line_vals,
            })

        if missed_rows:
            return self._download_missed_excel(missed_rows)

        if move:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'account.move',
                'res_id': move.id,
                'view_mode': 'form',
                'target': 'current',
            }

    def _download_missed_excel(self, missed_rows):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise UserError('openpyxl ბიბლიოთეკა არ არის დაინსტალირებული.')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ვერ მოიძებნა'

        headers = ['სახელი', 'საიდენტიფიკაციო კოდი', 'თანხა']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for row in missed_rows:
            ws.append([row['partner_name'], row['partner_vat'], row['amount']])

        stream = BytesIO()
        wb.save(stream)
        excel_b64 = base64.b64encode(stream.getvalue()).decode('utf-8')

        attachment = self.env['ir.attachment'].create({
            'name': 'pension_missed_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S'),
            'datas': excel_b64,
            'res_model': self._name,
            'res_id': 0,
            'type': 'binary',
        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'ნაწილობრივ შესრულდა',
                'message': '%d პარტნიორი ვერ მოიძებნა. ჩამოტვირთეთ სია.' % len(missed_rows),
                'type': 'warning',
                'sticky': True,
                'next': {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                },
            },
        }

    def _read_excel(self, data):
        try:
            import openpyxl
        except ImportError:
            raise UserError('openpyxl ბიბლიოთეკა არ არის დაინსტალირებული.')

        rows = []
        try:
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx < 8:
                    continue
                partner_name = row[1] if len(row) > 1 else None   # column B
                partner_vat = row[2] if len(row) > 2 else None    # column C
                amount_raw = row[4] if len(row) > 4 else None     # column E

                if not partner_name and not partner_vat:
                    continue

                try:
                    amount = float(amount_raw) if amount_raw not in (None, '') else 0.0
                except (ValueError, TypeError):
                    amount = 0.0

                if amount <= 0:
                    continue

                rows.append({
                    'partner_name': str(partner_name).strip() if partner_name else '',
                    'partner_vat': str(partner_vat).strip() if partner_vat else '',
                    'amount': amount,
                })
        except Exception as e:
            raise UserError('Excel ფაილის წაკითხვა ვერ მოხერხდა: %s' % e)
        return rows

    def _find_partner(self, name, vat):
        if vat:
            partner = self.env['res.partner'].search([('vat', '=', vat)], limit=1)
            if partner:
                return partner
        if name:
            partner = self.env['res.partner'].search([('name', '=', name)], limit=1)
            if partner:
                return partner
        return False
