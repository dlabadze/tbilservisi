# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)


class EmployeePhoneImportWizard(models.TransientModel):
    _name = 'employee.phone.import.wizard'
    _description = 'Employee Phone Import Wizard'

    excel_file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')

    def action_import(self):
        """Import employee phone numbers from Excel file"""
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))
        
        try:
            # Decode the file
            file_data = base64.b64decode(self.excel_file)
            
            # Detect file format and read accordingly
            filename_lower = (self.filename or '').lower()
            
            if filename_lower.endswith('.xlsx'):
                sheet_data = self._read_xlsx_file(file_data)
            elif filename_lower.endswith('.xls'):
                sheet_data = self._read_xls_file(file_data)
            else:
                raise UserError(_('Unsupported file format. Please upload .xls or .xlsx file.'))
            
            if not sheet_data or len(sheet_data) < 1:
                raise UserError(_('Excel file is empty.'))
            
            # Expected columns (0-based index):
            # 0: Employee (first column) - not used for search, just for reference
            # 1: identification_id (second column) - used to find employee
            # 2: Phone (third column) - phone number to set
            
            # Group rows by employee identification_id
            employee_phones = {}  # {employee_id: set of phone numbers}
            error_lines = []
            not_found_ids = set()  # Collect unique identification_ids that don't match any employee
            updated_count = 0
            
            for row_idx, row in enumerate(sheet_data, start=1):
                if not row or len(row) == 0:
                    continue
                
                try:
                    # Column 1 (second column): identification_id
                    identification_id = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    
                    if not identification_id:
                        error_lines.append(f'Row {row_idx}: Missing identification_id')
                        continue
                    
                    # Remove dot and everything after it from identification_id
                    if '.' in identification_id:
                        identification_id = identification_id.split('.')[0]
                    
                    # Find employee by identification_id
                    employee = self.env['hr.employee'].search([
                        ('identification_id', '=', identification_id)
                    ], limit=1)
                    
                    if not employee:
                        not_found_ids.add(identification_id)
                        error_lines.append(f'Row {row_idx}: Employee not found with identification_id: {identification_id}')
                        continue
                    
                    # Column 2 (third column): Phone number
                    phone = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    
                    if not phone:
                        error_lines.append(f'Row {row_idx}: Missing phone number')
                        continue
                    
                    # Remove dot and everything after it from phone
                    if '.' in phone:
                        phone = phone.split('.')[0]
                    
                    # Group by employee
                    if employee.id not in employee_phones:
                        employee_phones[employee.id] = set()
                    
                    employee_phones[employee.id].add(phone)
                    
                except Exception as e:
                    error_lines.append(f'Row {row_idx}: {str(e)}')
                    _logger.error(f'Error processing row {row_idx}: {str(e)}')
            
            # Update employees with phone numbers
            for employee_id, phone_set in employee_phones.items():
                employee = self.env['hr.employee'].browse(employee_id)
                
                # Get existing phones
                existing_phones = set()
                if employee.private_phone:
                    # Split by "; " if multiple phones exist
                    existing_phones = {p.strip() for p in employee.private_phone.split('; ') if p.strip()}
                
                # Merge with new phones
                all_phones = existing_phones | phone_set
                
                # Update employee.private_phone
                phone_string = '; '.join(sorted(all_phones))
                employee.write({'private_phone': phone_string})
                updated_count += 1
            
            # Build notification message
            message = f'Updated {updated_count} employee(s) with phone numbers.'
            
            # Add message for identification_ids not found
            if not_found_ids:
                ids_list = ', '.join(sorted(not_found_ids))
                message += f'\n\n[{ids_list}] "ამ პირად ნომრებზე თანამშრომელი ვერ მოიძებნა"'
            
            # Determine notification type
            notification_type = 'warning' if not_found_ids else 'success'
            
            # Return notification
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Complete'),
                    'message': message,
                    'type': notification_type,
                    'sticky': True if not_found_ids else False,
                }
            }
            
        except Exception as e:
            _logger.error(f'Error importing employee phones: {str(e)}')
            raise UserError(_('Error importing file: %s') % str(e))
    
    def _read_xlsx_file(self, file_data):
        """Read .xlsx file using openpyxl"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
            sheet = workbook.active
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            
            return data
            
        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        except Exception as e:
            raise UserError(_('Error reading .xlsx file: %s') % str(e))
    
    def _read_xls_file(self, file_data):
        """Read .xls file using xlrd"""
        try:
            import xlrd
            
            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)
            
            data = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    
                    if sheet.cell_type(row_idx, col_idx) == 3:  # XL_CELL_DATE
                        try:
                            cell_value = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                        except:
                            pass
                    
                    row.append(cell_value)
                data.append(row)
            
            return data
            
        except ImportError:
            raise UserError(_('xlrd library not installed. Please install: pip install xlrd'))
        except Exception as e:
            raise UserError(_('Error reading .xls file: %s') % str(e))

