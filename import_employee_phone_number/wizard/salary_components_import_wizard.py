import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import datetime
import logging

_logger = logging.getLogger(__name__)


class SalaryComponentsImportWizard(models.TransientModel):
    _name = 'salary.components.import.wizard'
    _description = 'Import Salary Components from Excel'

    excel_file = fields.Binary(string='Excel File', required=True, help='Upload Excel file with salary component data')
    filename = fields.Char(string='Filename')

    def action_import(self):
        """Import salary components from Excel file"""
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))

        try:
            # Decode the file
            file_data = base64.b64decode(self.excel_file)

            # Detect file format and read accordingly
            filename_lower = (self.filename or '').lower()

            if filename_lower.endswith('.xlsx'):
                # Read .xlsx file with openpyxl
                sheet_data = self._read_xlsx_file(file_data)
            elif filename_lower.endswith('.xls'):
                # Read .xls file with xlrd
                sheet_data = self._read_xls_file(file_data)
            else:
                raise UserError(_('Unsupported file format. Please upload .xls or .xlsx file.'))

            # Get header row
            if len(sheet_data) < 2:
                raise UserError(_('Excel file must have at least a header row and one data row.'))

            # Read data starting from row 1 (skip header)
            # Expected columns:
            # 0: თანამშრომელი - Employee name (informational only, skipped)
            # 1: პირადი ნომერი - identification_id
            # 2: დასახელება - x_studio_rule (hr.salary.rule name)
            # 3: დაწყების თარიღი - x_studio_start
            # 4: დასრულების თარიღი - x_studio_end
            # 5: თანხა - x_studio_compamount

            created_count = 0
            error_lines = []
            for row_idx in range(1, len(sheet_data)):
                try:
                    row = sheet_data[row_idx]
                    if not row or row == []:
                        continue

                    # Column 0: თანამშრომელი - Employee name (informational only, skip)

                    # Column 1: პირადი ნომერი - identification_id
                    identification_id = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    # Remove decimal point and everything after it (e.g., "123.0" -> "123")
                    if '.' in identification_id:
                        identification_id = identification_id.split('.')[0]

                    if not identification_id:
                        error_lines.append(f'Row {row_idx + 1}: Missing პირადი ნომერი (identification_id)')
                        continue

                    # Search for employee by identification_id
                    employee = self.env['hr.employee'].search([
                        ('identification_id', '=', identification_id)
                    ], limit=1)

                    if not employee:
                        error_lines.append(f'Row {row_idx + 1}: Employee not found with identification_id: {identification_id}')
                        continue

                    # Column 2: დასახელება - x_studio_rule (hr.salary.rule name)
                    comptypes = str(row[2]).strip() if len(row) > 2 and row[2] else ''

                    if not comptypes:
                        error_lines.append(f'Row {row_idx + 1}: Missing დასახელება (comptypes)')
                        continue

                    # Search for hr.salary.rule by name
                    salary_rule = self.env['hr.salary.rule'].search([
                        ('name', '=', comptypes)
                    ], limit=1)

                    if not salary_rule:
                        error_lines.append(f'Row {row_idx + 1}: Salary rule not found with name: {comptypes}')
                        continue

                    # Column 3: დაწყების თარიღი - x_studio_start
                    start_date = None
                    if len(row) > 3 and row[3]:
                        start_date = self._parse_date(row[3])

                    # Column 4: დასრულების თარიღი - x_studio_end
                    end_date = None
                    if len(row) > 4 and row[4]:
                        end_date = self._parse_date(row[4])

                    # Column 5: თანხა - x_studio_compamount
                    compamount = 0.0
                    if len(row) > 5 and row[5]:
                        try:
                            compamount = float(row[5])
                        except Exception:
                            compamount = 0.0

                    # Create salary component in employee's x_studio_salarytypes One2many field
                    vals = {
                        'x_studio_rule': salary_rule.id,
                        'x_studio_compamount': compamount,
                    }

                    if start_date:
                        vals['x_studio_start'] = start_date
                    if end_date:
                        vals['x_studio_end'] = end_date

                    # Add to employee's x_studio_salarytypes
                    employee.write({
                        'x_studio_salarytypes': [(0, 0, vals)]
                    })
                    created_count += 1

                except Exception as e:
                    error_lines.append(f'Row {row_idx + 1}: {str(e)}')
                    _logger.error(f'Error processing row {row_idx + 1}: {str(e)}')

            # Show result message - only success message
            if created_count > 0:
                message = f'Successfully imported {created_count} salary component(s).'
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Import Complete'),
                        'message': message,
                        'type': 'success',
                        'sticky': False,
                    }
                }
            else:
                # If no records were created, show an error message
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Import Failed'),
                        'message': _('No salary components were imported. Please check the Excel file format and data.'),
                        'type': 'danger',
                        'sticky': True,
                    }
                }

        except Exception as e:
            _logger.error(f'Error importing salary components: {str(e)}')
            raise UserError(_('Error importing file: %s') % str(e))

    def _read_xlsx_file(self, file_data):
        """Read .xlsx file using openpyxl"""
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
            sheet = workbook.active

            # Convert to list of lists
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))

            return data

        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        except Exception as e:
            raise UserError(_('Error reading .xlsx file: %s') % str(e))

    def _read_xls_file(self, file_data):
        """Read .xls file using xlrd"""
        try:
            import xlrd

            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)

            # Convert to list of lists
            data = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)

                    # Convert Excel dates
                    if sheet.cell_type(row_idx, col_idx) == 3:  # XL_CELL_DATE
                        try:
                            cell_value = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                        except Exception:
                            pass

                    row.append(cell_value)
                data.append(row)

            return data

        except ImportError:
            raise UserError(_('xlrd library not installed. Please install: pip install xlrd'))
        except Exception as e:
            raise UserError(_('Error reading .xls file: %s') % str(e))

    def _parse_date(self, date_value):
        """Parse date from various formats"""
        if not date_value:
            return None

        try:
            # If already a datetime object
            if isinstance(date_value, datetime):
                return date_value.date()

            # If it's a date object
            from datetime import date as date_type
            if isinstance(date_value, date_type):
                return date_value

            # If it's a string, try to parse
            if isinstance(date_value, str):
                date_str = date_value.strip()

                # Try different formats with dd/mm/yyyy and dd.mm.yyyy as primary formats
                date_formats = [
                    '%d/%m/%Y',      # 25/05/2025 (Primary format)
                    '%d.%m.%Y',      # 25.05.2025 (Primary format)
                    '%d-%m-%Y',      # 25-05-2025
                    '%Y-%m-%d',      # 2025-05-25
                    '%Y/%m/%d',      # 2025/05/25
                    '%m/%d/%Y',      # 05/25/2025
                    '%m-%d-%Y',      # 05-25-2025
                ]

                for fmt in date_formats:
                    try:
                        return datetime.strptime(date_str, fmt).date()
                    except Exception:
                        continue

                # Manual parsing with D/M/Y or D.M.Y as primary
                if '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) == 3:
                        try:
                            # Try D/M/Y format (Primary)
                            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                            from datetime import date as date_type
                            return date_type(year, month, day)
                        except Exception:
                            try:
                                # Try M/D/Y format
                                month, day, year = int(parts[0]), int(parts[1]), int(parts[2])
                                from datetime import date as date_type
                                return date_type(year, month, day)
                            except Exception:
                                pass
                elif '.' in date_str:
                    parts = date_str.split('.')
                    if len(parts) == 3:
                        try:
                            # Try D.M.Y format (Primary)
                            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                            from datetime import date as date_type
                            return date_type(year, month, day)
                        except Exception:
                            pass

            # If it's a number (Excel serial date)
            if isinstance(date_value, (int, float)):
                from datetime import timedelta
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=int(date_value))).date()

        except Exception as e:
            _logger.warning(f'Could not parse date: {date_value}, error: {str(e)}')

        return None
