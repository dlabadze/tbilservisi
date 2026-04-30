# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging


_logger = logging.getLogger(__name__)


class FitpassImportWizard(models.TransientModel):
    _name = 'fitpass.import.wizard'
    _description = 'Fitpass Import Wizard'

    start_row = fields.Integer(string='Start Row', default=2)
    end_row = fields.Integer(string='End Row', default=1000)
    excel_file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')

    def action_import(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))

        try:
            file_data = base64.b64decode(self.excel_file)
            filename_lower = (self.filename or '').lower()

            if filename_lower.endswith('.xlsx'):
                sheet_data = self._read_xlsx_file(file_data)
            elif filename_lower.endswith('.xls'):
                sheet_data = self._read_xls_file(file_data)
            else:
                raise UserError(_('Unsupported file format. Please upload .xls or .xlsx file.'))

            if not sheet_data or len(sheet_data) < 1:
                raise UserError(_('Excel file is empty.'))

            # Column 4 = identification_id (0-based index 3), Column 8 = amount (0-based index 7)
            data = []
            for row_idx in range(self.start_row - 1, min(self.end_row, len(sheet_data))):
                row = sheet_data[row_idx]
                if len(row) < 8:
                    continue
                identification_id = row[3]
                amount = row[7]
                if identification_id is None or identification_id == '':
                    continue
                identification_id = str(identification_id).strip()
                if '.' in identification_id:
                    identification_id = identification_id.split('.')[0]
                data.append({
                    'excel_row': row_idx + 1,
                    'identification_id': identification_id,
                    'amount': amount,
                })
            if not data:
                raise UserError(_('No valid rows found in the selected range.'))

            employees = self.env['hr.employee'].search([('identification_id', '!=', False)])
            updated_count = 0
            missed_rows = []
            for d in data:
                employee = employees.filtered(lambda e: str(e.identification_id or '').strip() == d['identification_id'])
                if not employee:
                    missed_rows.append({
                        'excel_row': d['excel_row'],
                        'identification_id': d['identification_id'],
                        'amount': d['amount'],
                        'error': _('Employee not found (identification_id)'),
                    })
                    continue
                employee = employee[:1]
                write_vals = {}
                if 'x_studio_fitpass' in employee._fields:
                    write_vals['x_studio_fitpass'] = d['amount']
                if write_vals:
                    employee.write(write_vals)
                updated_count += 1

            if missed_rows:
                notification = {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Import'),
                        'message': _('%s employee(s) updated. %s row(s) missed (no matching employee). Download the generated file for details.') % (updated_count, len(missed_rows)),
                        'type': 'warning',
                        'sticky': False,
                    },
                }
                attachment = self._generate_missed_rows_excel(missed_rows)
                download_action = {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                    'name': _('Download missed rows'),
                }
                if notification.get('params'):
                    notification['params']['next'] = download_action
                return notification
            else:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Import'),
                        'message': _('%s employee(s) updated.') % updated_count,
                        'type': 'success',
                        'sticky': False,
                    },
                }
        except Exception as e:
            _logger.error("Error importing Excel file: %s", e)
            raise UserError(_('Error parsing Excel file: %s') % str(e))

    def _read_xlsx_file(self, file_data):
        """Read .xlsx file using openpyxl"""
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
            sheet = workbook.active  # xlsx: use active sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            return data
        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        except Exception as e:
            raise UserError(_('Error reading .xlsx file: %s') % str(e))

    def _read_xls_file(self, file_data):
        """Read .xls file using xlrd"""
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)  # xls: use first sheet
            data = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if sheet.cell_type(row_idx, col_idx) == 3:  # XL_CELL_DATE
                        try:
                            cell_value = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                        except Exception:
                            pass
                    row.append(cell_value)
                data.append(row)
            return data
        except ImportError:
            raise UserError(_('xlrd library not installed. Please install: pip install xlrd'))
        except Exception as e:
            raise UserError(_('Error reading .xls file: %s') % str(e))

    def _generate_missed_rows_excel(self, missed_rows):
        """Generate an xlsx file with missed rows and return the ir.attachment record."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        wb = Workbook()
        ws = wb.active
        ws.title = 'Missed Rows'
        headers = [_('Excel Row'), _('Identification ID'), _('Amount'), _('Error')]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row_idx, row_data in enumerate(missed_rows, 2):
            ws.cell(row=row_idx, column=1, value=row_data.get('excel_row'))
            ws.cell(row=row_idx, column=2, value=row_data.get('identification_id'))
            ws.cell(row=row_idx, column=3, value=row_data.get('amount'))
            ws.cell(row=row_idx, column=4, value=row_data.get('error', ''))
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        attachment = self.env['ir.attachment'].create({
            'name': _('fitpass_import_missed_rows.xlsx'),
            'datas': base64.b64encode(buffer.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return attachment
