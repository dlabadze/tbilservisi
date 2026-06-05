# -*- coding: utf-8 -*-
from odoo import api, fields, models
from odoo.exceptions import UserError
import base64
from io import BytesIO
from datetime import datetime, date as date_cls
import logging
_logger = logging.getLogger(__name__)


class FuelExcelUploadWizard(models.TransientModel):
	_name = 'fuel.excel.upload.wizard'
	_description = 'Fuel Excel Upload Wizard'

	IMPORT_TYPE_TO_EXCEL_SHEET = {
		'sakutari': 'rented',
		'administratsia': 'own',
		'saerto': 'common',
	}

	date = fields.Date(
		string='თარიღი',
		required=True,
		default=fields.Date.context_today,
	)

	import_type = fields.Selection(
		selection=[
			('sakutari', 'იჯარით აღებული'),
			('administratsia', 'საკუთარი'),
			('saerto', 'საერთო'),
		],
		string='ტიპი',
		required=True,
	)

	file_data = fields.Binary(string='Excel ფაილი', required=True)
	file_name = fields.Char(string='ფაილის სახელი')

	def _get_excel_sheet_value(self):
		return self.IMPORT_TYPE_TO_EXCEL_SHEET.get(self.import_type)

	def _get_special_writeoff_location(self):
		location = self.env.ref('__export__.stock_location_881_5a9a8b7d', raise_if_not_found=False)
		if not location:
			raise UserError('სპეციალური ჩამოწერის ლოკაცია ვერ მოიძებნა.')
		return location

	def _get_special_product_template(self):
		return self.env.ref('__export__.product_template_24522_e6f6215f', raise_if_not_found=False)

	def _is_special_fuel_product(self, fuel_product_id):
		if not fuel_product_id:
			return False
		product = self.env['product.product'].browse(fuel_product_id)
		product_template = self._get_special_product_template()
		return bool(product.exists() and product_template and product.product_tmpl_id == product_template)

	def _get_department_writeoff_location(self, department):
		if department and 'x_studio_location' in department._fields and department.x_studio_location:
			return department.x_studio_location
		return False

	def _get_writeoff_location(self, department, fuel_product_id=False):
		if self.import_type in ('sakutari', 'administratsia'):
			return self._get_special_writeoff_location()
		if self.import_type == 'saerto' and self._is_special_fuel_product(fuel_product_id):
			return self._get_special_writeoff_location()
		return self._get_department_writeoff_location(department)

	def _normalize_vehicle_number(self, value):
		"""Normalize vehicle number by removing dashes and spaces.
		Example: 'aa-111-bb' or 'aa 111 bb' becomes 'aa111bb'
		"""
		if not value:
			return ''
		# Convert to string and remove dashes and spaces
		normalized = str(value).replace('-', '').replace(' ', '').strip().upper()
		return normalized

	def _parse_excel_date(self, value, datemode=None):
		"""Try to convert excel cell value to python date.
		- If datetime/date: return date
		- If float/int with datemode: xlrd serial to date
		- If string: try slicing YYYY-MM-DD from 'YYYY-MM-DD HH:MM:SS' then common formats
		Return None if cannot parse.
		"""
		if value is None or value == '':
			return None
		# Already date/datetime
		if isinstance(value, date_cls):
			return value
		if isinstance(value, datetime):
			return value.date()
		# Excel serial number (mostly in xls)
		try:
			from numbers import Number
			is_number = isinstance(value, Number)
		except Exception:
			is_number = False
		if is_number and datemode is not None:
			try:
				import xlrd  # type: ignore
				dt = xlrd.xldate_as_datetime(value, datemode)
				return dt.date()
			except Exception:
				pass
		# String parsing
		if isinstance(value, str):
			text = value.strip()
			# Slice YYYY-MM-DD if text like '2025-09-30 23:39:46'
			if len(text) >= 10 and text[0:4].isdigit() and text[5:7].isdigit() and text[8:10].isdigit():
				try:
					year = int(text[0:4])
					month = int(text[5:7])
					day = int(text[8:10])
					return date_cls(year, month, day)
				except Exception:
					pass
			for fmt in (
				"%Y-%m-%d",
				"%Y-%m-%d %H:%M:%S",
				"%d/%m/%Y",
				"%d-%m-%Y",
				"%d.%m.%Y",
				"%m/%d/%Y",
				"%d/%m/%y",
				"%m/%d/%y",
			):
				try:
					return datetime.strptime(text, fmt).date()
				except Exception:
					continue
		return None

	def _read_excel_sakutari(self, data: bytes) -> list:
		"""Return list of dicts with data for საკუთარი type.
		Excel read: active sheet only, same style as generate_sawvavi_wizard.
		Column B (index 1) = department_id (name or code)
		Date = wizard date
		Start reading from row 9
		"""
		rows = []
		stream = BytesIO(data)
		# Active sheet only (like generate_sawvavi_wizard): workbook.active, not by index
		try:
			import openpyxl  # type: ignore
			workbook = openpyxl.load_workbook(stream, data_only=True)
			sheet = workbook.active
			for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
				# Skip rows 1-8, start reading from row 9
				if idx < 10:
					continue
				# Column B is index 1 (0-based)
				department_val = row[1] if len(row) > 1 else None
				# Column G is index 6 (0-based)
				vehicle_num_val = row[6] if len(row) > 6 else None
				# Column K is index 10 (0-based) - gasoline amount
				gasoline_amount = row[10] if len(row) > 10 else None
				# Column L is index 11 (0-based) - diesel amount
				diesel_amount = row[11] if len(row) > 11 else None
				amount = gasoline_amount if gasoline_amount else diesel_amount
				if department_val:
					rows.append({
						'row_num': idx,
						'date': self.date,  # Use wizard date
						'department_name': str(department_val).strip(),
						'vehicle_number': str(vehicle_num_val).strip() if vehicle_num_val else '',
						'amount': float(amount) if amount not in (None, '') else 0.0,
						'raw_row': row,  # Store full row for additional processing
					})
			return rows
		except Exception as e:
			raise UserError('ვერ წავიკითხე Excel ფაილი. გთხოვთ ატვირთოთ .xlsx ფაილი.\n%s' % e)

	def _read_excel_administratsia(self, data: bytes) -> list:
		"""Return list of dicts with data for ადმინისტრაციის მანქანები type.
		Excel read: active sheet only, same style as generate_sawvavi_wizard.
		Column A (index 0) = vehicle number
		Column F (index 5) = department_id
		Column I (index 8) = start_odometer
		Column J (index 9) = odometer
		Column K (index 10) = monthly_mileage
		Column L (index 11) = start_balance
		Column M (index 12) = filled_qty
		Column N (index 13) = consumed_qty
		Column O (index 14) = final_balance
		Date = wizard date
		Start reading from row 3
		"""
		rows = []
		stream = BytesIO(data)
		# Active sheet only (like generate_sawvavi_wizard): workbook.active, not by index
		try:
			import openpyxl  # type: ignore
			workbook = openpyxl.load_workbook(stream, data_only=True)
			sheet = workbook.active
			for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
				# Skip rows 1-2, start reading from row 3
				if idx < 3:
					continue
				# Column A is index 0 (0-based) - vehicle number
				vehicle_num_val = row[0] if len(row) > 0 else None
				# Column F is index 5 (0-based)
				department_val = row[5] if len(row) > 5 else None
				# Column I is index 8 (0-based) - start_odometer
				start_odometer = row[8] if len(row) > 8 else None
				# Column J is index 9 (0-based) - odometer
				odometer = row[9] if len(row) > 9 else None
				# Column K is index 10 (0-based) - monthly_mileage
				monthly_mileage = row[10] if len(row) > 10 else None
				# Column L is index 11 (0-based) - start_balance
				start_balance = row[11] if len(row) > 11 else None
				# Column M is index 12 (0-based) - filled_qty
				filled_qty = row[12] if len(row) > 12 else None
				# Column N is index 13 (0-based) - consumed_qty
				consumed_qty = row[13] if len(row) > 13 else None
				# Column O is index 14 (0-based) - final_balance
				final_balance = row[14] if len(row) > 14 else None
				if department_val or vehicle_num_val:
					rows.append({
						'row_num': idx,
						'date': self.date,  # Use wizard date
						'department_name': str(department_val).strip() if department_val else '',
						'vehicle_number': str(vehicle_num_val).strip() if vehicle_num_val else '',
						'start_odometer': float(start_odometer) if start_odometer not in (None, '') else 0.0,
						'odometer': float(odometer) if odometer not in (None, '') else 0.0,
						'monthly_mileage': float(monthly_mileage) if monthly_mileage not in (None, '') else 0.0,
						'start_balance': float(start_balance) if start_balance not in (None, '') else 0.0,
						'filled_qty': float(filled_qty) if filled_qty not in (None, '') else 0.0,
						'consumed_qty': float(consumed_qty) if consumed_qty not in (None, '') else 0.0,
						'final_balance': float(final_balance) if final_balance not in (None, '') else 0.0,
						'raw_row': row,  # Store full row for additional processing
					})
			return rows
		except Exception as e:
			raise UserError('ვერ წავიკითხე Excel ფაილი. გთხოვთ ატვირთოთ .xlsx ფაილი.\n%s' % e)

	def _read_excel_saerto(self, data: bytes) -> list:
		"""Return list of dicts with data for საერთო type.
		Excel read: active sheet only, same style as generate_sawvavi_wizard.
		Column B (index 1) = department_id
		Column E (index 4) = vehicle number (search using x_studio_related_field_5g9_1j6vebbdt)
		Column F (index 5) = start_odometer (shifted from E)
		Column G (index 6) = start_worked_hours (shifted from F)
		Column H (index 7) = start_balance (gasoline) (shifted from G)
		Column I (index 8) = start_balance (diesel) (shifted from H)
		Column J (index 9) = other_received (gasoline) (shifted from I)
		Column K (index 10) = other_received (diesel) (shifted from J)
		Column L (index 11) = other_transferred (gasoline) (shifted from K)
		Column M (index 12) = other_transferred (diesel) (shifted from L)
		Column N (index 13) = odometer (shifted from M)
		Column O (index 14) = monthly_mileage (shifted from N)
		Column P (index 15) = worked_hours (shifted from O)
		Column Q (index 16) = filled_qty (gasoline) (shifted from P)
		Column R (index 17) = filled_qty (diesel) (shifted from Q)
		Column S (index 18) = final_balance (gasoline) (shifted from R)
		Column T (index 19) = final_balance (diesel) (shifted from S)
		Date = wizard date
		Start reading from row 10
		"""
		rows = []
		stream = BytesIO(data)
		# Only openpyxl, active sheet (like generate_sawvavi_wizard)
		try:
			import openpyxl  # type: ignore
		except ImportError:
			raise UserError('openpyxl ბიბლიოთეკა არ არის დაინსტალირებული.')
		try:
			workbook = openpyxl.load_workbook(stream, data_only=True)
			sheet = workbook.active
			for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
				# Skip rows 1-9, start reading from row 10
				if idx < 10:
					continue
				# Column B is index 1 (0-based)
				department_val = row[1] if len(row) > 1 else None
				# Column E is index 4 (0-based) - vehicle number (changed from D)
				vehicle_num_val = row[4] if len(row) > 4 else None
				# Column F is index 5 (0-based) - start_odometer (shifted from E)
				start_odometer = row[5] if len(row) > 5 else None
				# Column G is index 6 (0-based) - start_worked_hours (shifted from F)
				start_worked_hours = row[6] if len(row) > 6 else None
				# Column H is index 7 (0-based) - start_balance_gasoline (shifted from G)
				start_balance_gasoline = row[7] if len(row) > 7 else None
				# Column I is index 8 (0-based) - start_balance_diesel (shifted from H)
				start_balance_diesel = row[8] if len(row) > 8 else None
				# Column J is index 9 (0-based) - other_received_gasoline (shifted from I)
				other_received_gasoline = row[9] if len(row) > 9 else None
				# Column K is index 10 (0-based) - other_received_diesel (shifted from J)
				other_received_diesel = row[10] if len(row) > 10 else None
				# Column L is index 11 (0-based) - other_transferred_gasoline (shifted from K)
				other_transferred_gasoline = row[11] if len(row) > 11 else None
				# Column M is index 12 (0-based) - other_transferred_diesel (shifted from L)
				other_transferred_diesel = row[12] if len(row) > 12 else None
				# Column N is index 13 (0-based) - odometer (shifted from M)
				odometer = row[13] if len(row) > 13 else None
				# Column O is index 14 (0-based) - monthly_mileage (shifted from N)
				monthly_mileage = row[14] if len(row) > 14 else None
				# Column P is index 15 (0-based) - worked_hours (shifted from O)
				worked_hours = row[15] if len(row) > 15 else None
				# Column Q is index 16 (0-based) - filled_qty_gasoline (shifted from P)
				filled_qty_gasoline = row[16] if len(row) > 16 else None
				# Column R is index 17 (0-based) - filled_qty_diesel (shifted from Q)
				filled_qty_diesel = row[17] if len(row) > 17 else None
				# Column S is index 18 (0-based) - final_balance_gasoline (shifted from R)
				final_balance_gasoline = row[18] if len(row) > 18 else None
				# Column T is index 19 (0-based) - final_balance_diesel (shifted from S)
				final_balance_diesel = row[19] if len(row) > 19 else None

				start_amount = start_balance_gasoline if start_balance_gasoline else start_balance_diesel
				other_received = other_received_gasoline if other_received_gasoline else other_received_diesel
				other_transferred = other_transferred_gasoline if other_transferred_gasoline else other_transferred_diesel
				filled_qty = filled_qty_gasoline if filled_qty_gasoline else filled_qty_diesel
				final_balance = final_balance_gasoline if final_balance_gasoline else final_balance_diesel

				
				# Helper function to safely convert to float
				def safe_float(value):
					if value in (None, ''):
						return 0.0
					try:
						return float(value)
					except (ValueError, TypeError):
						return 0.0
				
				# Skip rows where numeric columns contain text (like headers)
				if department_val or vehicle_num_val:
					# Try to process the row, skip if conversion fails (header rows, etc.)
					try:
						rows.append({
							'row_num': idx,
							'date': self.date,  # Use wizard date
							'department_name': str(department_val).strip() if department_val else '',
							'vehicle_number': str(vehicle_num_val).strip() if vehicle_num_val else '',
							'start_odometer': safe_float(start_odometer),
							'start_worked_hours': safe_float(start_worked_hours),
							'start_balance': safe_float(start_amount),
							'other_received': safe_float(other_received),
							'other_transferred': safe_float(other_transferred),
							'odometer': safe_float(odometer),
							'monthly_mileage': safe_float(monthly_mileage),
							'worked_hours': safe_float(worked_hours),
							'filled_qty': safe_float(filled_qty),
							'final_balance': safe_float(final_balance),
							'raw_row': row,  # Store full row for additional processing
						})
					except Exception:
						continue
			return rows
		except Exception as e:
			raise UserError('ვერ წავიკითხე Excel ფაილი. გთხოვთ ატვირთოთ .xlsx ფაილი.\n%s' % e)

	def action_confirm(self):
		self.ensure_one()
		if not self.file_data:
			raise UserError('ატვირთეთ Excel ფაილი.')

		data = base64.b64decode(self.file_data)
		
		# Read Excel based on import type
		if self.import_type == 'sakutari':
			# _logger.info(f'Reading Excel for იჯარით აღებული type')
			rows = self._read_excel_sakutari(data)
			# _logger.info(f'Rows: {rows}')
		elif self.import_type == 'administratsia':
			rows = self._read_excel_administratsia(data)
		elif self.import_type == 'saerto':
			rows = self._read_excel_saerto(data)
		else:
			raise UserError('არასწორი ტიპი.')

		if not rows:
			raise UserError('ფაილში მონაცემები ვერ მოიძებნა.')

		# Process rows and create fuel.management records
		if self.import_type == 'sakutari':
			return self._process_sakutari_rows(rows)
		elif self.import_type == 'administratsia':
			return self._process_administratsia_rows(rows)
		elif self.import_type == 'saerto':
			return self._process_saerto_rows(rows)
		
		# Success - return action to refresh the fuel.management list
		action = self.env.ref('fuel_management.action_fuel_management', raise_if_not_found=False)
		if action:
			return action.read()[0]
		return {'type': 'ir.actions.act_window_close'}

	def _process_sakutari_rows(self, rows):
		"""Process rows for საკუთარი type and create fuel.management records."""
		to_create = []
		missed_rows = []
		
		for r in rows:
			# Prepare tracking variables for validation
			dept_name = r.get('department_name', '').strip()
			vehicle_number = r.get('vehicle_number', '').strip()
			vehicle_fuel_type = ''
			employee_name = ''
			location_name = ''
			missed_reason = []
			
			# Track what was found for missed excel display
			dept_found = False
			vehicle_found = False
			fuel_type_found = False
			employee_found = False
			location_found = False
			
			# Validate department
			department = self.env['hr.department'].search([('name', '=', dept_name)], limit=1)
			if not department:
				missed_reason.append('დეპარტამენტი ვერ მოიძებნა')
			else:
				dept_found = True
			
			# Find and validate vehicle
			vehicle_id = False
			fuel_product_id = False
			employee_id = False
			location_id = False
			filled_qty = 0.0
			consumed_qty = 0.0
			
			vehicle = None
			if vehicle_number:
				normalized_vehicle_num = self._normalize_vehicle_number(vehicle_number)
				if normalized_vehicle_num:
					vehicle = self.env['fleet.vehicle'].search([
						('x_studio_carnumbexp', '=', normalized_vehicle_num)
					], limit=1)
					
					if not vehicle:
						missed_reason.append('მანქანა ვერ მოიძებნა')
					else:
						vehicle_id = vehicle.id
						vehicle_found = True
						
						# Validate and get fuel product
						if hasattr(vehicle, 'x_studio_product') and vehicle.x_studio_product:
							product_template = vehicle.x_studio_product
							if product_template.product_variant_ids:
								fuel_product_id = product_template.product_variant_ids[0].id
								vehicle_fuel_type = product_template.name or ''
								fuel_type_found = True
							else:
								missed_reason.append('საწვავის ტიპი ვერ მოიძებნა')
						else:
							missed_reason.append('საწვავის ტიპი ვერ მოიძებნა')
						
						# Validate and get employee (commented: computed on fuel.management from vehicle_id log_drivers)
						# if 'log_drivers' in vehicle._fields and vehicle.log_drivers:
						# 	log_drivers = vehicle.log_drivers.sorted(key=lambda r: r.create_date or r.id, reverse=True)
						# 	if log_drivers:
						# 		newest_log = log_drivers[0]
						# 		if hasattr(newest_log, 'driver_employee_id') and newest_log.driver_employee_id:
						# 			employee_id = newest_log.driver_employee_id.id
						# 			employee_name = newest_log.driver_employee_id.name or ''
						# 			employee_found = True
						# 		else:
						# 			missed_reason.append('თანამშრომელი ვერ მოიძებნა')
						# 	else:
						# 		missed_reason.append('თანამშრომელი ვერ მოიძებნა')
						# else:
						# 	missed_reason.append('თანამშრომელი ვერ მოიძებნა')
						
						# Rented sheet records always use the configured special write-off location.
						location = self._get_writeoff_location(department, fuel_product_id)
						if location:
							location_id = location.id
							location_name = location.display_name or location.name or ''
							location_found = True
						else:
							missed_reason.append('ლოკაცია ვერ მოიძებნა')
						
						# Get amount based on vehicle fuel_type
						amount = r.get('amount', 0.0)
						filled_qty = amount
						consumed_qty = amount
			else:
				missed_reason.append('მანქანა ვერ მოიძებნა')
			
			# If any validation failed, add to missed rows
			if missed_reason or not department:
				# Show 'ვერ მოიძებნა' for fields that were not found
				dept_display = dept_name if dept_found else 'ვერ მოიძებნა'
				vehicle_display = vehicle_number if vehicle_found else 'ვერ მოიძებნა'
				fuel_type_display = vehicle_fuel_type if fuel_type_found else 'ვერ მოიძებნა'
				employee_display = employee_name if employee_found else 'ვერ მოიძებნა'
				location_display = location_name if location_found else 'ვერ მოიძებნა'
				
				missed_rows.append({
					'row_num': r.get('row_num'),
					'department_name': dept_display,
					'vehicle_number': vehicle_display,
					'fuel_type': fuel_type_display,
					'employee_name': employee_display,
					'location_name': location_display,
					'error': ', '.join(missed_reason) if missed_reason else 'დეპარტამენტი ვერ მოიძებნა'
				})
				continue
			
			# Create record values (employee_id, department_id, parent_department_id computed on fuel.management from vehicle_id log_drivers)
			# parent_department = department
			# while parent_department.parent_id:
			# 	parent_department = parent_department.parent_id
			vals = {
				'excel_sheet': self._get_excel_sheet_value(),
				'date': r.get('date'),
				# 'parent_department_id': parent_department.id,
				# 'department_id': department.id,
				'vehicle_id': vehicle_id,
				'fuel_product_id': fuel_product_id,
				# 'employee_id': employee_id,
				'writeoff_location_id': location_id,
				'filled_qty': filled_qty,
				'consumed_qty': consumed_qty,
			}
			to_create.append(vals)
		
		# Check if there are any missed rows - if yes, do not create any records
		if missed_rows:
			return self._generate_and_download_missed_excel(missed_rows, 0)
		
		# Create all records only if all rows are valid (no missed rows)
		created_count = 0
		if to_create:
			self.env['fuel.management'].create(to_create)
			created_count = len(to_create)
		
		# Success - return action to refresh the fuel.management list
		action = self.env.ref('fuel_management.action_fuel_management', raise_if_not_found=False)
		if action:
			return action.read()[0]
		return {'type': 'ir.actions.act_window_close'}
	
	def _process_administratsia_rows(self, rows):
		"""Process rows for ადმინისტრაციის მანქანები type and create fuel.management records."""
		to_create = []
		missed_rows = []
		
		for r in rows:
			# Get department and vehicle info
			dept_name = r.get('department_name', '').strip()
			vehicle_number = r.get('vehicle_number', '').strip()
			vehicle_fuel_type = ''
			location_name = ''
			missed_reason = []
			
			# Track what was found for missed excel display
			dept_found = False
			vehicle_found = False
			fuel_type_found = False
			location_found = False
			
			# Skip row if both department and vehicle are missing
			if not dept_name and not vehicle_number:
				continue
			
			# Find department
			department = None
			if dept_name:
				department = self.env['hr.department'].search([('name', '=', dept_name)], limit=1)
				if not department:
					missed_reason.append('დეპარტამენტი ვერ მოიძებნა')
				else:
					dept_found = True
			
			# Find vehicle
			vehicle_id = False
			fuel_product_id = False
			location_id = False
			vehicle = None
			
			if vehicle_number:
				# For administratsia type, search using x_studio_related_field_5g9_1j6vebbdt field
				vehicle_number_str = str(vehicle_number).strip()
				if vehicle_number_str:
					vehicle = self.env['fleet.vehicle'].search([
						('vin_sn', '=', vehicle_number_str)
					], limit=1)
					if not vehicle:
						missed_reason.append('მანქანა ვერ მოიძებნა')
					else:
						vehicle_id = vehicle.id
						vehicle_found = True
						
						# Get fuel product from vehicle's x_studio_product
						if hasattr(vehicle, 'x_studio_product') and vehicle.x_studio_product:
							product_template = vehicle.x_studio_product
							if product_template.product_variant_ids:
								fuel_product_id = product_template.product_variant_ids[0].id
								vehicle_fuel_type = product_template.name or ''
								fuel_type_found = True
							else:
								missed_reason.append('საწვავის ტიპი ვერ მოიძებნა')
						else:
							missed_reason.append('საწვავის ტიპი ვერ მოიძებნა')
						
						# Own sheet records use the department write-off location.
						location = self._get_writeoff_location(department, fuel_product_id)
						if location:
							location_id = location.id
							location_name = location.display_name or location.name or ''
							location_found = True
						else:
							missed_reason.append('ლოკაცია ვერ მოიძებნა')
				else:
					# Vehicle number is empty
					missed_reason.append('მანქანა ვერ მოიძებნა')
			
			# If vehicle or department not found, add to missed rows
			if missed_reason:
				# Show 'ვერ მოიძებნა' for fields that were not found
				dept_display = dept_name if dept_found else 'ვერ მოიძებნა'
				vehicle_display = vehicle_number if vehicle_found else 'ვერ მოიძებნა'
				fuel_type_display = vehicle_fuel_type if fuel_type_found else 'ვერ მოიძებნა'
				location_display = location_name if location_found else 'ვერ მოიძებნა'
				
				missed_rows.append({
					'row_num': r.get('row_num'),
					'department_name': dept_display,
					'vehicle_number': vehicle_display,
					'fuel_type': fuel_type_display,
					'location_name': location_display,
				})
				continue
			
			# Create record values
			vals = {
				'excel_sheet': self._get_excel_sheet_value(),
				'date': r.get('date'),
				'start_odometer': r.get('start_odometer', 0.0),
				'odometer': r.get('odometer', 0.0),
				'monthly_mileage': r.get('monthly_mileage', 0.0),
				'start_balance': r.get('start_balance', 0.0),
				'filled_qty': r.get('filled_qty', 0.0),
				'consumed_qty': r.get('consumed_qty', 0.0),
				'final_balance': r.get('final_balance', 0.0),
				'writeoff_location_id': location_id,
			}
			
			# Add department fields if department found (commented: computed on fuel.management from vehicle_id log_drivers)
			# if department:
			# 	parent_department = department
			# 	while parent_department.parent_id:
			# 		parent_department = parent_department.parent_id
			# 	vals['department_id'] = department.id
			# 	vals['parent_department_id'] = parent_department.id
			
			# Add vehicle-related fields
			if vehicle_id:
				vals['vehicle_id'] = vehicle_id
			if fuel_product_id:
				vals['fuel_product_id'] = fuel_product_id
			
			to_create.append(vals)
		
		# Check if there are any missed rows - if yes, do not create any records
		if missed_rows:
			return self._generate_and_download_missed_excel_administratsia(missed_rows, 0)
		
		# Create all records only if all rows are valid (no missed rows)
		created_count = 0
		if to_create:
			self.env['fuel.management'].create(to_create)
			created_count = len(to_create)
		
		# Success - return action to refresh the fuel.management list
		action = self.env.ref('fuel_management.action_fuel_management', raise_if_not_found=False)
		if action:
			return action.read()[0]
		return {'type': 'ir.actions.act_window_close'}
	
	def _process_saerto_rows(self, rows):
		"""Process rows for საერთო type and create fuel.management records."""
		to_create = []
		missed_rows = []
		
		for r in rows:
			# Get department and vehicle info
			dept_name = r.get('department_name', '').strip()
			vehicle_number = r.get('vehicle_number', '').strip()
			vehicle_fuel_type = ''
			location_name = ''
			missed_reason = []
			
			# Track what was found for missed excel display
			dept_found = False
			vehicle_found = False
			fuel_type_found = False
			location_found = False
			
			# Skip row if both department and vehicle are missing
			if not dept_name and not vehicle_number:
				continue
			
			# Find department
			department = None
			if dept_name:
				department = self.env['hr.department'].search([('name', '=', dept_name)], limit=1)
				if not department:
					missed_reason.append('დეპარტამენტი ვერ მოიძებნა')
				else:
					dept_found = True
			
			# Find vehicle
			vehicle_id = False
			fuel_product_id = False
			location_id = False
			vehicle = None
			fuel_type = None
			
			if vehicle_number:
				# For საერთო type, search using x_studio_related_field_5g9_1j6vebbdt field
				vehicle_number_str = str(vehicle_number).strip()
				if vehicle_number_str:
					vehicle = self.env['fleet.vehicle'].search([
						('vin_sn', '=', vehicle_number_str)
					], limit=1)
					
					if not vehicle:
						missed_reason.append('მანქანა ვერ მოიძებნა')
					else:
						vehicle_id = vehicle.id
						vehicle_found = True
						
						# Get fuel product from vehicle's x_studio_product
						if hasattr(vehicle, 'x_studio_product') and vehicle.x_studio_product:
							product_template = vehicle.x_studio_product
							if product_template.product_variant_ids:
								fuel_product_id = product_template.product_variant_ids[0].id
								vehicle_fuel_type = product_template.name or ''
								fuel_type_found = True
							else:
								missed_reason.append('საწვავის ტიპი ვერ მოიძებნა')
						else:
							missed_reason.append('საწვავის ტიპი ვერ მოიძებნა')
						
						# Get fuel_type from vehicle
						if hasattr(vehicle, 'fuel_type'):
							fuel_type = vehicle.fuel_type
						
						# Common sheet records use the special location for the configured fuel,
						# otherwise they fall back to the department write-off location.
						location = self._get_writeoff_location(department, fuel_product_id)
						if location:
							location_id = location.id
							location_name = location.display_name or location.name or ''
							location_found = True
						else:
							missed_reason.append('ლოკაცია ვერ მოიძებნა')
				else:
					# Normalized vehicle number is empty
					missed_reason.append('მანქანა ვერ მოიძებნა')
			else:
				# No vehicle number provided
				missed_reason.append('მანქანა ვერ მოიძებნა')
			
			# CRITICAL: Skip row if vehicle, fuel_type, or location not found
			# Check all required fields
			if not vehicle_id:
				if 'მანქანა ვერ მოიძებნა' not in missed_reason:
					missed_reason.append('მანქანა ვერ მოიძებნა')
			if not fuel_product_id:
				if 'საწვავის ტიპი ვერ მოიძებნა' not in missed_reason:
					missed_reason.append('საწვავის ტიპი ვერ მოიძებნა')
			if not location_id:
				if 'ლოკაცია ვერ მოიძებნა' not in missed_reason:
					missed_reason.append('ლოკაცია ვერ მოიძებნა')
			
			# If any required field is missing, skip this row
			if missed_reason:
				# Show 'ვერ მოიძებნა' for fields that were not found
				dept_display = dept_name if dept_found else 'ვერ მოიძებნა'
				vehicle_display = vehicle_number if vehicle_found else 'ვერ მოიძებნა'
				fuel_type_display = vehicle_fuel_type if fuel_type_found else 'ვერ მოიძებნა'
				location_display = location_name if location_found else 'ვერ მოიძებნა'
				
				missed_rows.append({
					'row_num': r.get('row_num'),
					'department_name': dept_display,
					'vehicle_number': vehicle_display,
					'fuel_type': fuel_type_display,
					'location_name': location_display,
				})
				continue
			
			# Determine which columns to use based on fuel_type
			start_balance = 0.0
			other_received = 0.0
			other_transferred = 0.0
			filled_qty = 0.0
			final_balance = 0.0
			
			start_balance = r.get('start_balance', 0.0)
			other_received = r.get('other_received', 0.0)
			other_transferred = r.get('other_transferred', 0.0)
			filled_qty = r.get('filled_qty', 0.0)
			final_balance = r.get('final_balance', 0.0)
		
			
			# Create record values
			vals = {
				'excel_sheet': self._get_excel_sheet_value(),
				'date': r.get('date'),
				'start_odometer': r.get('start_odometer', 0.0),
				'start_worked_hours': r.get('start_worked_hours', 0.0),
				'odometer': r.get('odometer', 0.0),
				'monthly_mileage': r.get('monthly_mileage', 0.0),
				'worked_hours': r.get('worked_hours', 0.0),
				'start_balance': start_balance,
				'other_received': other_received,
				'other_transferred': other_transferred,
				'filled_qty': filled_qty,
				'final_balance': final_balance,
				'writeoff_location_id': location_id,
			}
			
			# Add department fields if department found (commented: computed on fuel.management from vehicle_id log_drivers)
			# if department:
			# 	parent_department = department
			# 	while parent_department.parent_id:
			# 		parent_department = parent_department.parent_id
			# 	vals['department_id'] = department.id
			# 	vals['parent_department_id'] = parent_department.id
			
			# Add vehicle-related fields
			if vehicle_id:
				vals['vehicle_id'] = vehicle_id
			if fuel_product_id:
				vals['fuel_product_id'] = fuel_product_id
			
			to_create.append(vals)
		
		# Check if there are any missed rows - if yes, do not create any records
		if missed_rows:
			return self._generate_and_download_missed_excel_saerto(missed_rows, 0)
		
		# Create all records only if all rows are valid (no missed rows)
		created_count = 0
		if to_create:
			self.env['fuel.management'].create(to_create)
			created_count = len(to_create)
		
		# Success - return action to refresh the fuel.management list
		action = self.env.ref('fuel_management.action_fuel_management', raise_if_not_found=False)
		if action:
			return action.read()[0]
		return {'type': 'ir.actions.act_window_close'}
	
	def _generate_and_download_missed_excel(self, missed_rows, created_count):
		"""Generate Excel file with missed rows and return download action (for საკუთარი type)."""
		try:
			import openpyxl
			from openpyxl.styles import Font, PatternFill
			
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = 'დაკარგული ჩანაწერები'
			
			# Header row
			headers = ['რიგი', 'სამსახური', 'მანქანა', 'საწვავის ტიპი', 'თანამშრომელი', 'ლოკაცია']
			ws.append(headers)
			
			# Style header
			for cell in ws[1]:
				cell.font = Font(bold=True)
				cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
			
			# Add missed rows
			for mr in missed_rows:
				ws.append([
					mr.get('row_num'),
					mr.get('department_name'),
					mr.get('vehicle_number'),
					mr.get('fuel_type'),
					mr.get('employee_name'),
					mr.get('location_name')
				])
			
			# Save to binary
			stream = BytesIO()
			wb.save(stream)
			excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')
			
			# Generate filename
			filename = 'დაკარგული_ჩანაწერები_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
			
			# Show message and download
			if created_count == 0:
				message = 'არცერთი ჩანაწერი არ შეიქმნა!\n%s ჩანაწერი ვერ დამუშავდა და ყველა ჩანაწერი უარყოფილია.' % len(missed_rows)
				title = 'ატვირთვა ვერ მოხერხდა'
				msg_type = 'danger'
			else:
				message = '%s ჩანაწერი შეიქმნა წარმატებით.\n%s ჩანაწერი ვერ დამუშავდა.' % (
					created_count, len(missed_rows)
				)
				title = 'ატვირთვა დასრულდა'
				msg_type = 'warning'
			
			# Create temporary attachment for download
			attachment = self.env['ir.attachment'].create({
				'name': filename,
				'datas': excel_data,
				'res_model': 'fuel.excel.upload.wizard',
				'res_id': 0,
				'type': 'binary',
			})
			
			# Return download action with message
			return {
				'type': 'ir.actions.client',
				'tag': 'display_notification',
				'params': {
					'title': title,
					'message': message,
					'type': msg_type,
					'sticky': False,
					'next': {
						'type': 'ir.actions.act_url',
						'url': '/web/content/%s?download=true' % attachment.id,
						'target': 'self',
					}
				}
			}
			
		except Exception as e:
			raise UserError('ვერ შეიქმნა შეცდომების ფაილი: %s' % str(e))
	
	def _generate_and_download_missed_excel_administratsia(self, missed_rows, created_count):
		"""Generate Excel file with missed rows for ადმინისტრაციის მანქანები type."""
		try:
			import openpyxl
			from openpyxl.styles import Font, PatternFill
			
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = 'დაკარგული ჩანაწერები'
			
			# Header row - five columns: row, department, vehicle, fuel type, location
			headers = ['რიგი', 'სამსახური', 'მანქანა', 'საწვავის ტიპი', 'ლოკაცია']
			ws.append(headers)
			
			# Style header
			for cell in ws[1]:
				cell.font = Font(bold=True)
				cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
			
			# Add missed rows
			for mr in missed_rows:
				ws.append([
					mr.get('row_num'),
					mr.get('department_name'),
					mr.get('vehicle_number'),
					mr.get('fuel_type'),
					mr.get('location_name')
				])
			
			# Save to binary
			stream = BytesIO()
			wb.save(stream)
			excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')
			
			# Generate filename
			filename = 'დაკარგული_ჩანაწერები_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
			
			# Show message and download
			if created_count == 0:
				message = 'არცერთი ჩანაწერი არ შეიქმნა!\n%s ჩანაწერი ვერ დამუშავდა და ყველა ჩანაწერი უარყოფილია.' % len(missed_rows)
				title = 'ატვირთვა ვერ მოხერხდა'
				msg_type = 'danger'
			else:
				message = '%s ჩანაწერი შეიქმნა წარმატებით.\n%s ჩანაწერი ვერ დამუშავდა.' % (
					created_count, len(missed_rows)
				)
				title = 'ატვირთვა დასრულდა'
				msg_type = 'warning'
			
			# Create temporary attachment for download
			attachment = self.env['ir.attachment'].create({
				'name': filename,
				'datas': excel_data,
				'res_model': 'fuel.excel.upload.wizard',
				'res_id': 0,
				'type': 'binary',
			})
			
			# Return download action with message
			return {
				'type': 'ir.actions.client',
				'tag': 'display_notification',
				'params': {
					'title': title,
					'message': message,
					'type': msg_type,
					'sticky': False,
					'next': {
						'type': 'ir.actions.act_url',
						'url': '/web/content/%s?download=true' % attachment.id,
						'target': 'self',
					}
				}
			}
			
		except Exception as e:
			raise UserError('ვერ შეიქმნა შეცდომების ფაილი: %s' % str(e))
	
	def _generate_and_download_missed_excel_saerto(self, missed_rows, created_count):
		"""Generate Excel file with missed rows for საერთო type."""
		try:
			import openpyxl
			from openpyxl.styles import Font, PatternFill
			
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = 'დაკარგული ჩანაწერები'
			
			# Header row - five columns: row, department, vehicle, fuel type, location
			headers = ['რიგი', 'სამსახური', 'მანქანა', 'საწვავის ტიპი', 'ლოკაცია']
			ws.append(headers)
			
			# Style header
			for cell in ws[1]:
				cell.font = Font(bold=True)
				cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
			
			# Add missed rows
			for mr in missed_rows:
				ws.append([
					mr.get('row_num'),
					mr.get('department_name'),
					mr.get('vehicle_number'),
					mr.get('fuel_type'),
					mr.get('location_name')
				])
			
			# Save to binary
			stream = BytesIO()
			wb.save(stream)
			excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')
			
			# Generate filename
			filename = 'დაკარგული_ჩანაწერები_%s.xlsx' % fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
			
			# Show message and download
			if created_count == 0:
				message = 'არცერთი ჩანაწერი არ შეიქმნა!\n%s ჩანაწერი ვერ დამუშავდა და ყველა ჩანაწერი უარყოფილია.' % len(missed_rows)
				title = 'ატვირთვა ვერ მოხერხდა'
				msg_type = 'danger'
			else:
				message = '%s ჩანაწერი შეიქმნა წარმატებით.\n%s ჩანაწერი ვერ დამუშავდა.' % (
					created_count, len(missed_rows)
				)
				title = 'ატვირთვა დასრულდა'
				msg_type = 'warning'
			
			# Create temporary attachment for download
			attachment = self.env['ir.attachment'].create({
				'name': filename,
				'datas': excel_data,
				'res_model': 'fuel.excel.upload.wizard',
				'res_id': 0,
				'type': 'binary',
			})
			
			# Return download action with message
			return {
				'type': 'ir.actions.client',
				'tag': 'display_notification',
				'params': {
					'title': title,
					'message': message,
					'type': msg_type,
					'sticky': False,
					'next': {
						'type': 'ir.actions.act_url',
						'url': '/web/content/%s?download=true' % attachment.id,
						'target': 'self',
					}
				}
			}
			
		except Exception as e:
			raise UserError('ვერ შეიქმნა შეცდომების ფაილი: %s' % str(e))