from odoo import models, fields, api
import base64
import io
from odoo.exceptions import UserError
from openpyxl import load_workbook
from datetime import datetime

class WizardInvoiceImport(models.TransientModel):
    _name = 'wizard.invoice.import'
    _description = 'Invoice Import Wizard'

    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    journal_name_ = fields.Selection([('fizikuri', 'სამშენებლო მასალების განთავსება (ფიზიკური პირი)'),
                                  ('iuridiuli', 'სამშენებლო მასალების განთავსება (იურიდიული პირი)')],
                                 string="ჟურნალის დასახელება:", default='fizikuri')

    def action_import_invoices(self):
        if not self.file:
            raise UserError("Please upload an Excel file")

        # Decode Excel file
        data = base64.b64decode(self.file)
        workbook = load_workbook(filename=io.BytesIO(data), data_only=True)
        sheet = workbook.active

        journal_name_map = {
            'fizikuri': "სამშენებლო მასალების განთავსება (ფიზიკური პირი)",
            'iuridiuli': "სამშენებლო მასალების განთავსება (იურიდიული პირი)",
        }
        journal_name = journal_name_map.get(self.journal_name_)
        journal = self.env['account.journal'].search([('name', '=', journal_name)], limit=1)
        if not journal:
            raise UserError(f"Journal not found: {journal_name}")

        product = self.env['product.product'].search([('name', '=', 'სამშენებლო მასალები')], limit=1)
        if not product:
            raise UserError("Product 'სამშენებლო მასალები' not found")

        account = self.env['account.account'].search([('code', '=', '6110')], limit=1)
        if not account:
            raise UserError("Account 6110 Product Sales not found")

        # Read headers (row 2)
        headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        col_index = {header: idx for idx, header in enumerate(headers)}

        # Iterate over data rows (starting row 3)
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row[col_index['კლიენტი']]:
                continue  # skip empty rows

            partner_name = str(row[col_index['კლიენტი']]).strip()
            partner = self.env['res.partner'].search([('name', '=', partner_name)], limit=1)
            if not partner:
                partner = self.env['res.partner'].create({'name': partner_name})

            # Compose basis string
            basis = f": {row[col_index['მანქანა']]} / {row[col_index['მძღოლი']]} / {row[col_index['ტვირთამწეობა']]}"

            # Check if invoice exists (same partner + contract + date)

            invoice_date = row[col_index['შექმნის თარიღი']]
            invoice_year = invoice_date.year if invoice_date else datetime.today().year
            prefix = 'სმგფ' if self.journal_name_ == 'fizikuri' else 'სმგი'
            last_invoice = self.env['account.move'].search([
                ('move_type', '=', 'out_invoice'),
                ('name', 'like', f'{prefix}/{invoice_year}/%')
            ], order='id desc', limit=1)

            if last_invoice and last_invoice.name:
                try:
                    last_index = int(last_invoice.name.split('/')[-1])
                except:
                    last_index = 0
            else:
                last_index = 0
            new_index = last_index + 1
            invoice_name = f"{prefix}/{invoice_year}/{str(new_index).zfill(3)}"

            self.env['account.move'].create({
                    'name': invoice_name,
                    'move_type': 'out_invoice',
                    'partner_id': partner.id,
                    'journal_id': journal.id,
                    'invoice_date': row[col_index['შექმნის თარიღი']],
                    'identification_code': row[col_index['საიდენტიფიკაციო']],
                    'basis': basis,
                    'contract_num': row[col_index['კონტრაქტი']],
                    'invoice_line_ids': [(0, 0, {
                        'product_id': product.id,
                        'quantity': 1.0,
                        'price_unit': row[col_index['ღირებულება']],
                        'account_id': account.id,
                    })]
                })
