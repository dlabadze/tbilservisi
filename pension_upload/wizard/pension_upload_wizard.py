from odoo import models, fields, _, api
from odoo.exceptions import UserError
import base64
import io

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

class PensionUploadWizard(models.TransientModel):
    _name = 'pension.upload.wizard'
    _description = 'Pension Upload Wizard'

    file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')

    def action_upload(self):
        if not self.file:
            raise UserError(_("Please upload a file."))
        
        if not load_workbook:
            raise UserError(_("The 'openpyxl' library is missing."))

        try:
            file_data = base64.b64decode(self.file)
            workbook = load_workbook(io.BytesIO(file_data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid file! Please upload a valid Excel file.\nError: %s") % str(e))

        not_found_employees = []
        updated_count = 0
        
        # Skip header row (min_row=2)
        # Using values_only=True returns tuples of cell values
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for index, row in enumerate(rows, start=2):
            # Check for empty rows/cells to avoid errors
            if not row or len(row) < 3: 
                continue

            # Column 2 (index 1) -> Personal Number
            personal_number = row[1]
            # Column 3 (index 2) -> Pension Status
            pension_status = row[2]

            if not personal_number:
                continue

            # Ensure personal_number is string
            personal_number = str(personal_number).strip()
            if isinstance(pension_status, str):
                 pension_status = pension_status.strip()

            employee = self.env['hr.employee'].search([('identification_id', '=', personal_number)], limit=1)

            if employee:
                if pension_status == 'კი':
                    employee.x_studio_pension = True
                    updated_count += 1
            else:
                not_found_employees.append(personal_number)

        message = _("Pension status updated for %s employees.") % updated_count
        if not_found_employees:
            message += "\n\n" + _("The following personal numbers were not found:\n%s") % ", ".join(not_found_employees)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Upload Completed"),
                'message': message,
                'type': 'warning' if not_found_employees else 'success',
                'sticky': True if not_found_employees else False,
            }
        }
