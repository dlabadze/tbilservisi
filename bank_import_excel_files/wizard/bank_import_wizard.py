import base64
import io
import openpyxl
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import datetime, date
from openpyxl.utils.datetime import from_excel


class BankImportWizard(models.TransientModel):
    _name = "bank.import.wizard"
    _description = "Bank Import Excel Wizard"

    # Journal selection
    journal_type = fields.Selection([
        ('ge90', 'GE90LB0113172770216000GEL'),
        ('tbc', 'თიბისი ფეი'),
        ('ge17', 'GE17TB7722636020100004GEL'),
        ('ge81', 'GE81BG0000000499211307GEL'),
        ('ge40', 'GE40BG0000000499210105GEL'),
        ('ge22', 'GE22TB7722636020100011GEL'),
        ('ge56', 'GE56TB7722645067800007GEL'),
        ('ge09', 'GE09LB0113150423521000GEL'),
        ('ge21', 'GE21LB0113122202198000GEL'),
        ('ge73', 'GE73BG0000000176620200GEL'),
    ], string="აირჩიეთ ჟურნალი", required=True)

    line_from = fields.Integer(string='პირველი ხაზი')
    line_to = fields.Integer(string='ბოლო ხაზი')
    excel_file = fields.Binary(string="აირჩიეთ Excel ფაილი", required=True)
    file_name = fields.Char(string="ფაილის სახელი")

    def action_import_excel(self):
        self = self.sudo()
        self.ensure_one()
        if not self.excel_file:
            raise UserError("გთხოვთ ატვირთოთ Excel ფაილი.")

        JOURNAL_MAPPING = {
            'tbc': 'თიბისი ფეი',
            'ge17': 'GE17TB7722636020100004GEL',
            'ge90': 'GE90LB0113172770216000GEL',
            'ge81': 'GE81BG0000000499211307GEL',
            'ge40': 'GE40BG0000000499210105GEL',
            'ge22': 'GE22TB7722636020100011GEL',
            'ge56': 'GE56TB7722645067800007GEL',
            'ge09': 'GE09LB0113150423521000GEL',
            'ge21': 'GE21LB0113122202198000GEL',
            'ge73': 'GE73BG0000000176620200GEL',
        }

        journal_name = JOURNAL_MAPPING.get(self.journal_type)
        journal = self.env['account.journal'].search([('name', '=', journal_name)], limit=1)
        if not journal:
            raise UserError(_("ჟურნალი ვერ მოიძებნა: %s") % journal_name)
        file_content = base64.b64decode(self.excel_file)

        if not file_content:
            raise ValueError("Decoded file content is empty.")
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True, read_only=True)
            sheet = wb.active
        except Exception as e:
            print(f"Error loading workbook: {e}")
        account_1242 = self.env['account.account'].search([('code', '=', '1242')], limit=1)
        statement_lines = self.env['account.bank.statement.line']

        def parse_number(value):
            if value is None:
                return 0.0
            try:
                s = str(value).replace('\xa0', '').replace(',', '').strip()
                if s == '':
                    return 0.0
                return float(s)
            except Exception:
                return 0.0

        # ---------------------------
        # Existing TBC logic
        # ---------------------------
        if self.journal_type == 'tbc':
            header_row_index = None
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row and 'id' in [str(c).lower() if c else '' for c in row]:
                    header_row_index = i
                    break
            if header_row_index is None:
                raise UserError("Couldn't find the header row (expected a row with 'id')")

            headers = [str(cell).strip() if cell else '' for cell in
                       next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))]
            col_index = {header: idx for idx, header in enumerate(headers)}

            min_row = self.line_from or (header_row_index + 1)
            max_row = self.line_to if self.line_to else sheet.max_row

            rows = []
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
                if not row or not row[col_index.get('id', -1)]:
                    continue
                rows.append(row)

            suspense_acc_1242 = self.env['account.account'].search([('code', '=', '1210.01.01')], limit=1)

            if not suspense_acc_1242:
                raise UserError("სისტემაში ანგარიში კოდით '1210.01.01' ვერ მოიძებნა! გთხოვთ შექმნათ ანგარიშთა გეგმაში.")

            vat_to_name_map = {}
            for r in rows:
                idx_vat = col_index.get('მომხმარებლის იდენთიფიკატორი')
                idx_name = col_index.get('მომხმარებელი')
                vat_val = str(r[idx_vat]).strip() if idx_vat is not None else ''
                name_val = str(r[idx_name]).strip() if idx_name is not None else ''
                if vat_val and vat_val != 'None':
                    vat_to_name_map[vat_val] = name_val

            existing_partners = self.env['res.partner'].search([('vat', 'in', list(vat_to_name_map.keys()))])
            partner_map = {p.vat: p for p in existing_partners}

            missing_vats = set(vat_to_name_map.keys()) - set(partner_map.keys())
            if missing_vats:
                new_partners_vals = [{'name': vat_to_name_map[v], 'vat': v} for v in missing_vats]
                if new_partners_vals:
                    new_partners = self.env['res.partner'].create(new_partners_vals)
                    partner_map.update({p.vat: p for p in new_partners})

            line_vals_list = []
            for row in rows:
                row_vat = str(row[col_index.get('მომხმარებლის იდენთიფიკატორი')]).strip()
                partner = partner_map.get(row_vat)

                danishnuleba = f":{row[col_index.get('ხელშეკრ. ნომერი')]} / {row[col_index.get('ხელშეკრულების პერიოდი')]}"
                amount = parse_number(row[col_index.get('თანხა\n(თეთრი)')])
                raw_date_value = row[col_index.get('ჩარიცხვის დრო')]
                parsed_date = self.parse_excel_date(raw_date_value)
                tx_id_idx = col_index.get('Transaction ID')
                name = f"TBC/{row[tx_id_idx]}" if tx_id_idx is not None else 'TBC/import'

                client_type = row[col_index.get('კლიენტის ტიპი')] if col_index.get('კლიენტის ტიპი') is not None else None

                partner_id = partner.id if partner else False

                if partner and partner.category_id and partner.category_id.name == "ფიზიკური":
                    target_account_id = self.env['account.account'].search([('code', '=', '1400.01.1421')], limit=1)
                else:
                    target_account_id = self.env['account.account'].search([('code', '=', '1400.01.1422')], limit=1)

                if not target_account_id:
                    raise UserError("ანგარიში 1400.01.1421 ან 1400.01.1422 ვერ მოიძებნა. გთხოვთ შეამოწმოთ ანგარიშთა გეგმა.")

                line_vals_list.append({
                    'name': name,
                    'move_type': 'entry',
                    'partner_id': partner_id,
                    'date': parsed_date,
                    'identification_code': row_vat,
                    'journal_id': journal.id,
                    'danishnuleba': danishnuleba,
                    'invoice_line_ids': [
                        (0, 0, {
                            'account_id': journal.default_account_id.id,
                            'partner_id': partner_id,
                            'debit': amount / 100 if amount > 0 else 0,
                            'credit': abs(amount / 100) if amount < 0 else 0,
                        }),
                        (0, 0, {
                            'account_id': target_account_id.id,
                            'partner_id': partner_id,
                            'debit': abs(amount / 100) if amount < 0 else 0,
                            'credit': amount / 100 if amount > 0 else 0,
                        })
                    ]
                })

                if partner:
                    partner.write({
                        'company_type': 'person' if client_type in ['ფ/პ'] else 'company'
                    })

            statement_lines = self.env['account.move'].create(line_vals_list)

        # ---------------------------
        # GE17 / GE22 / GE56 Logic
        # ---------------------------
        elif self.journal_type in ['ge17', 'ge22', 'ge56']:
            sheet = wb[wb.sheetnames[1]]
            header_row_index = None
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row and 'პარტნიორი' in [str(c) for c in row if c]:
                    header_row_index = i
                    break
            if header_row_index is None:
                header_row_index = 1

            headers = [str(c).strip() if c else '' for c in
                       next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))]
            col_index = {h: i for i, h in enumerate(headers)}

            min_row = self.line_from or (header_row_index + 1)
            max_row = self.line_to or sheet.max_row

            rows = []
            vat_to_name_map = {}

            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
                if not row or not row[col_index.get('დანიშნულება', -1)]:
                    continue

                row_vat = str(row[col_index.get('პარტნიორის საგადასახადო კოდი')]).strip() if col_index.get(
                    'პარტნიორის საგადასახადო კოდი') else ''
                row_name = str(row[col_index.get('პარტნიორი')]).strip()

                if row_vat:
                    vat_to_name_map[row_vat] = row_name

                rows.append(row)

            Partner = self.env['res.partner']
            existing_partners = Partner.search([('vat', 'in', list(vat_to_name_map.keys()))])
            partner_map = {p.vat: p for p in existing_partners}

            missing_vats = set(vat_to_name_map.keys()) - set(partner_map.keys())

            if missing_vats:
                new_partners_vals = []
                for vat in missing_vats:
                    new_partners_vals.append({
                        'name': vat_to_name_map[vat],
                        'vat': vat
                    })
                if new_partners_vals:
                    new_partners = Partner.create(new_partners_vals)
                    partner_map.update({p.vat: p for p in new_partners})

            line_vals_list = []
            for row in rows:
                row_vat = str(row[col_index.get('პარტნიორის საგადასახადო კოდი')]).strip() if col_index.get(
                    'პარტნიორის საგადასახადო კოდი') else ''
                pname_idx = col_index.get('პარტნიორი')
                pname = str(row[pname_idx]).strip() if pname_idx is not None else ''

                partner = partner_map.get(row_vat)

                danishnuleba = f": {row[col_index.get('დანიშნულება')]} / {row[col_index.get('დამატებითი ინფორმაცია')]}"

                incoming = row[col_index.get('შემოსული თანხა')] if col_index.get('შემოსული თანხა') else 0
                outgoing = row[col_index.get('გასული თანხა')] if col_index.get('გასული თანხა') else 0

                # Ensure numbers are parsed correctly (assuming parse_number handles None/Empty)
                amount_in = parse_number(incoming) if incoming else 0.0
                amount_out = parse_number(outgoing) if outgoing else 0.0
                amount = amount_in or -amount_out

                date = self.parse_excel_date(row[col_index.get('თარიღი')])
                if not date:
                    continue

                line_vals = {
                    'name': danishnuleba,
                    'date': date,
                    'amount': amount,
                    'journal_id': journal.id,
                    'payment_ref': danishnuleba,
                    'saidentifikacio_code': row_vat,
                }

                # Apply partner info
                if partner:
                    # Update partner details as in original code
                    partner.write({
                        'vat': row_vat,  # Ensures consistency
                        'company_type': 'company' if any(x in pname for x in ['შპს', 'ი/მ', 'სს']) else 'person'
                    })

                    if amount < 0:
                        line_vals['partner_id'] = partner.id

                line_vals_list.append(line_vals)

            statement_lines = self.env['account.bank.statement.line'].create(line_vals_list)

        # ---------------------------
        # LIBERTY Logic (GE90, GE09, GE21)
        # ---------------------------
        elif self.journal_type in ['ge90', 'ge09', 'ge21']:
            sheet = wb[wb.sheetnames[1]]
            header_row_index = None
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                low_cells = [str(c).strip() if c else '' for c in row]
                if 'თარიღი' in low_cells or 'Date' in low_cells:
                    header_row_index = i
                    break
            if header_row_index is None:
                header_row_index = 2

            headers = [str(c).strip() if c else '' for c in
                       next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))]
            col_index = {h: i for i, h in enumerate(headers)}

            min_row = self.line_from or (header_row_index + 1)
            max_row = self.line_to or sheet.max_row

            rows = []
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
                if not any(row):
                    continue
                rows.append(row)


            vat_to_name_map = {}

            for row in rows:
                row_vat = row[col_index.get('პარტნიორის საგადასახადო კოდი')]
                row_name = row[col_index.get('პარტნიორი')]

                in_val = parse_number(row[col_index.get('შემოსული თანხა', 0)]) if 'შემოსული თანხა' in col_index else 0
                out_val = parse_number(row[col_index.get('გასული თანხა', 0)]) if 'გასული თანხა' in col_index else 0

                if (in_val or out_val) and row_vat:
                    vat_str = str(row_vat).strip()
                    name_str = str(row_name).strip() if row_name else ''
                    if vat_str:
                        vat_to_name_map[vat_str] = name_str

            Partner = self.env['res.partner']

            existing_partners = Partner.search([('vat', 'in', list(vat_to_name_map.keys()))])

            partner_map = {p.vat: p for p in existing_partners}

            missing_vats = set(vat_to_name_map.keys()) - set(partner_map.keys())
            if missing_vats:
                new_partners_vals = []
                for vat in missing_vats:
                    new_partners_vals.append({
                        'name': vat_to_name_map[vat] or vat,  # Fallback to VAT as name if name is empty
                        'vat': vat
                    })
                if new_partners_vals:
                    new_partners = Partner.create(new_partners_vals)
                    partner_map.update({p.vat: p for p in new_partners})

            line_vals_list = []
            for row in rows:
                desc = row[col_index.get('დანიშნულება', '')] or ''
                add_info = row[col_index.get('დამატებითი ინფორმაცია', '')] or ''
                payment_ref = f"{desc} / {add_info}".strip(' / ')
                idx_date = col_index.get('თარიღი')
                date_val = self.parse_excel_date(row[idx_date]) if idx_date is not None else False

                in_val = parse_number(row[col_index.get('შემოსული თანხა', 0)]) if 'შემოსული თანხა' in col_index else 0
                out_val = parse_number(row[col_index.get('გასული თანხა', 0)]) if 'გასული თანხა' in col_index else 0

                if in_val == 0 and out_val == 0:
                    continue

                amount = -out_val or in_val

                row_vat_raw = row[col_index.get('პარტნიორის საგადასახადო კოდი')]
                row_vat = str(row_vat_raw).strip() if row_vat_raw else ''
                pname = str(row[col_index.get('პარტნიორი', '')]).strip()

                partner = partner_map.get(row_vat)

                line_vals_list.append({
                    'name': payment_ref,
                    'date': date_val or False,
                    'amount': amount,
                    'journal_id': journal.id,
                    'payment_ref': payment_ref,
                    'saidentifikacio_code': row_vat,
                })

                if partner:
                    partner.write({
                        'vat': row_vat,
                        'company_type': 'company' if any(x in pname for x in ['შპს', 'ი/მ', 'სს']) else 'person'
                    })

                    if amount < 0:
                        line_vals_list[-1]['partner_id'] = partner.id

            statement_lines = self.env['account.bank.statement.line'].create(line_vals_list)

        # ---------------------------
        # GE81 / GE40 / GE73 Logic
        # ---------------------------
        elif self.journal_type in ['ge81', 'ge40', 'ge73']:
            sheet = wb.active

            BG_COLUMNS = {
                'თარიღი': 0,
                'დებეტი': 3,
                'კრედიტი': 4,
                'ოპერაციის შინაარსი': 5,
                'გამგზავნის დასახელება': 9,
                'გამგზავნის საიდენტიფიკაციო კოდი': 10,
                'გამგზავნი ბანკის დასახელება': 13,
                'მიმღების დასახელება': 14,
                'მიმღების საიდენტიფიკაციო კოდი': 15,
            }

            header_row = None
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                low_cells = [str(c).strip() if c else '' for c in row]
                if 'თარიღი' in low_cells or 'Date' in low_cells:
                    header_row = i
                    break

            if header_row is not None:
                headers = [str(c).strip() if c else '' for c in
                           next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
                col_index = {h: i for i, h in enumerate(headers)}
                min_row = self.line_from or (header_row + 1)
            else:
                col_index = BG_COLUMNS
                data_start = 1
                for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if any(row):
                        data_start = i
                        break
                min_row = self.line_from or data_start

            max_row = self.line_to or sheet.max_row
            rows = [r for r in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True) if any(r)]

            vat_to_name_map = {}

            for row in rows:
                idx_debit = col_index.get('დებეტი')
                debit = parse_number(row[idx_debit]) if idx_debit is not None else 0

                if debit:
                    sid = str(row[col_index.get('მიმღების საიდენტიფიკაციო კოდი')]).strip()
                    pname = str(row[col_index.get('მიმღების დასახელება')]).strip()

                    if sid:
                        vat_to_name_map[sid] = pname

            Partner = self.env['res.partner']

            existing_partners = Partner.search([('vat', 'in', list(vat_to_name_map.keys()))])
            partner_map = {p.vat: p for p in existing_partners}

            missing_vats = set(vat_to_name_map.keys()) - set(partner_map.keys())
            if missing_vats:
                new_partners_vals = []
                for vat in missing_vats:
                    new_partners_vals.append({
                        'name': vat_to_name_map[vat] or vat,
                        'vat': vat
                    })
                if new_partners_vals:
                    new_partners = Partner.create(new_partners_vals)
                    partner_map.update({p.vat: p for p in new_partners})
            # -----------------------------------------------------------------------

            line_vals_list = []
            for row in rows:
                idx_debit = col_index.get('დებეტი')
                idx_credit = col_index.get('კრედიტი')
                debit = parse_number(row[idx_debit]) if idx_debit is not None else 0
                credit = parse_number(row[idx_credit]) if idx_credit is not None else 0
                payment_ref = str(row[col_index.get('ოპერაციის შინაარსი')]) or ''

                if credit:
                    amount = credit
                    pname = row[col_index.get('გამგზავნი ბანკის დასახელება')]
                    sid = row[col_index.get('გამგზავნის საიდენტიფიკაციო კოდი')]
                elif debit:
                    amount = -debit
                    pname = row[col_index.get('მიმღების დასახელება')]
                    sid = row[col_index.get('მიმღების საიდენტიფიკაციო კოდი')]
                else:
                    continue

                pname = str(pname).strip() if pname else ''
                sid = str(sid).strip() if sid else ''

                partner = partner_map.get(sid)

                line_vals_list.append({
                    'date': self.parse_excel_date(row[col_index.get('თარიღი')]) if col_index.get('თარიღი') is not None else False,
                    'amount': amount,
                    'journal_id': journal.id,
                    'payment_ref': payment_ref,
                    'name': payment_ref or pname,
                    'saidentifikacio_code': sid,
                })

                if partner:
                    partner.write({
                        'vat': sid,
                        'company_type': 'company' if any(x in pname for x in ['შპს', 'ი/მ', 'სს']) else 'person'
                    })


                if amount < 0:
                    line_vals_list[-1]['partner_id'] = partner.id if partner else False

            statement_lines = self.env['account.bank.statement.line'].create(line_vals_list)

        # ---------------------------
        # Fallback
        # ---------------------------
        else:
            raise UserError("იმპორტის ლოგიკა დაემატება მოგვიანებით.")

        if not statement_lines:
            raise UserError(_("არ შეიქმნა არცერთი ჩანაწერი. გთხოვთ გადაამოწმოთ ფაილი."))

        # Return Action
        if self.journal_type == 'tbc':
            return {
                "type": "ir.actions.act_window",
                "name": "Bank Statement Lines",
                "res_model": "account.move",
                "view_mode": "list,form",
                "domain": [("id", "in", statement_lines.ids)],
            }

        return {
            "type": "ir.actions.act_window",
            "name": "Bank Statement Lines",
            "res_model": "account.bank.statement.line",
            "view_mode": "list,form",
            "domain": [("id", "in", statement_lines.ids)],
        }

    @staticmethod
    def parse_excel_date(value):
        if not value:
            return False
        if isinstance(value, (datetime, date)):
            return value.date() if isinstance(value, datetime) else value
        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return False
        if isinstance(value, str):
            value = " ".join(value.strip().split())
            for fmt in (
                    "%d-%m-%Y %H:%M:%S",
                    "%Y-%m-%d %H:%M:%S",
                    "%d/%m/%Y %H:%M:%S",
                    "%Y/%m/%d %H:%M:%S",
                    "%d-%m-%Y",
                    "%Y-%m-%d",
                    "%d/%m/%Y",
                    "%m/%d/%Y",
                    "%d.%m.%Y",
                    "%d.%m.%Y %H:%M:%S",
            ):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue
        return False