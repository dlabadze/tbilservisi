import base64
from collections import defaultdict
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.exceptions import UserError


ATTENDANCE_X = 'X'
ATTENDANCE_D = 'D'
ATTENDANCE_G = 'G'
ATTENDANCE_FIELD = 'x_studio_selection_field_99n_1j76jab36'

VIRTUAL_ATT_X = '__att_x__'
VIRTUAL_ATT_D = '__att_d__'
VIRTUAL_ATT_G = '__att_g__'

STATE_LABELS = {
    'draft':  'Draft',
    'verify': 'Waiting',
    'done':   'Done',
    'paid':   'Paid',
    'cancel': 'Canceled',
}

NUMERIC_TYPES  = {'integer', 'float', 'monetary'}
MONEY_TYPES    = {'monetary', 'float'}

DEFAULT_COLUMNS = [
    ('number','Reference',14),
    ('employee_id','Employee',28),
    ('department_id','Department',22),
    ('job_id','Job Position',22),
    ('payslip_run_id','Batch',18),
    ('date_from','From',12),
    ('date_to','To',12),
    ('basic_wage','Basic Wage',14),
    ('gross_wage','Gross Wage',14),
    ('net_wage','Net Wage',14),
    ('state','Status',12),
    (VIRTUAL_ATT_X,'დასწრება (X)', 14),
    (VIRTUAL_ATT_D, 'დასვენება (D)',14),
    (VIRTUAL_ATT_G,'გაცდენა (G)',16),
]

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
DATA_FONT = Font(name='Calibri', size=10)
ALT_FILL = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='D6E4BC', end_color='D6E4BC', fill_type='solid')
TOTAL_FONT = Font(bold=True, name='Calibri', size=10)
TITLE_FONT = Font(bold=True, name='Calibri', size=13, color='1F4E79')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right',  vertical='center')
THIN_SIDE = Side(style='thin', color='B0B0B0')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


class PayslipExportColumn(models.TransientModel):
    _name = 'payslip.export.column'
    _description = 'Payslip Export Column'
    _order = 'sequence, id'

    wizard_id = fields.Many2one(
        'payslip.excel.export.wizard',
        ondelete='cascade',
        required=True,
    )
    sequence = fields.Integer(default=10)

    field_name = fields.Selection(
        selection='_get_payslip_fields',
        string='Field',
    )
    virtual_key = fields.Selection([
        (VIRTUAL_ATT_X, 'დასწრება (X)'),
        (VIRTUAL_ATT_D, 'დასვენება (D)'),
        (VIRTUAL_ATT_G, 'გაცდენა (G)'),
    ], string='Virtual Column')

    label = fields.Char(
        string='Column Label',
        required=True,
        help='Header shown in the Excel file. Auto-filled when you pick a field.',
    )
    col_width = fields.Integer(string='Width', default=16)

    @api.model
    def _get_payslip_fields(self):
        fields_records = self.env['ir.model.fields'].sudo().search([
            ('model', '=', 'hr.payslip'),
            ('ttype', 'not in', ['one2many', 'many2many', 'binary', 'html'])
        ])
        return [(f.name, f.field_description) for f in fields_records]

    @api.onchange('field_name')
    def _onchange_field_name(self):
        if self.field_name:
            field = self.env['ir.model.fields'].sudo().search([
                ('model', '=', 'hr.payslip'),
                ('name', '=', self.field_name)
            ], limit=1)
            if field:
                self.label = field.field_description
                ttype = field.ttype
                self.col_width = 22 if ttype in ('many2one', 'char', 'text') else 14

    @api.onchange('virtual_key')
    def _onchange_virtual_key(self):
        if self.virtual_key:
            mapping = dict(self._fields['virtual_key'].selection)
            self.label = mapping.get(self.virtual_key, self.virtual_key)
            self.col_width = 14

    def _get_cell_value(self, slip, att_counts):

        if self.virtual_key == VIRTUAL_ATT_X:
            return att_counts[slip.employee_id.id][ATTENDANCE_X]
        if self.virtual_key == VIRTUAL_ATT_D:
            return att_counts[slip.employee_id.id][ATTENDANCE_D]
        if self.virtual_key == VIRTUAL_ATT_G:
            return att_counts[slip.employee_id.id][ATTENDANCE_G]

        if not self.field_name:
            return ''

        fname = self.field_name
        if fname not in slip._fields:
            return ''

        ttype = slip._fields[fname].type
        val = slip[fname]

        if ttype == 'many2one':
            return val.name if val else ''
        if ttype == 'selection':
            if fname == 'state':
                return STATE_LABELS.get(val, val) if val else ''
            sel_dict = dict(slip._fields[fname].selection if not callable(
                slip._fields[fname].selection) else slip._fields[fname].selection(slip))
            return sel_dict.get(val, val) if val else ''
        if ttype == 'boolean':
            return 'Yes' if val else 'No'
        if ttype == 'date':
            return val.strftime('%d.%m.%Y') if val else ''
        if val is False and ttype not in NUMERIC_TYPES:
            return ''
        return val

    def _is_numeric(self):
        if self.virtual_key in (VIRTUAL_ATT_X, VIRTUAL_ATT_D, VIRTUAL_ATT_G):
            return True
        if self.field_name and self.field_name in self.env['hr.payslip']._fields:
            return self.env['hr.payslip']._fields[self.field_name].type in NUMERIC_TYPES
        return False

    def _is_money(self):
        if self.field_name and self.field_name in self.env['hr.payslip']._fields:
            return self.env['hr.payslip']._fields[self.field_name].type in MONEY_TYPES
        return False

class PayslipExcelExportWizard(models.TransientModel):
    _name = 'payslip.excel.export.wizard'
    _description = 'Payslip Excel Export Wizard'

    date_from= fields.Date(string='From', required=True)
    date_to= fields.Date(string='To',   required=True)
    department_id = fields.Many2one(
        'hr.department',
        string='Department',
        help=(
            'Leave empty → all departments.\n'
            'Department with NO parent → it + all children.\n'
            'Department WITH a parent → only that department.'
        ),
    )
    column_ids = fields.One2many(
        'payslip.export.column',
        'wizard_id',
        string='Columns',
    )

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if 'column_ids' in fields_list:
            config = self.env['payslip.export.config'].get_config()

            if config.column_ids:
                col_lines = []
                for saved in config.column_ids.sorted('sequence'):
                    col_lines.append((0, 0, {
                        'sequence':    saved.sequence,
                        'field_name':  saved.field_name or False,
                        'virtual_key': saved.virtual_key or False,
                        'label':       saved.label,
                        'col_width':   saved.col_width,
                    }))
                res['column_ids'] = col_lines
            else:
                col_lines = []
                for seq, (key, label, width) in enumerate(DEFAULT_COLUMNS, start=1):
                    line = {
                        'sequence':    seq * 10,
                        'label':       label,
                        'col_width':   width,
                        'field_name':  False,
                        'virtual_key': False,
                    }
                    if key in (VIRTUAL_ATT_X, VIRTUAL_ATT_D, VIRTUAL_ATT_G):
                        line['virtual_key'] = key
                    else:
                        line['field_name'] = key
                    col_lines.append((0, 0, line))
                res['column_ids'] = col_lines
        return res

    def _get_child_department_ids(self, department):
        ids = [department.id]
        for child in self.env['hr.department'].search([('parent_id', '=', department.id)]):
            ids.extend(self._get_child_department_ids(child))
        return ids

    def _get_department_ids(self):
        if not self.department_id:
            return None
        if not self.department_id.parent_id:
            return self._get_child_department_ids(self.department_id)
        return [self.department_id.id]

    def _get_attendance_counts(self, employee_ids, date_from, date_to):
        counts = defaultdict(lambda: {ATTENDANCE_X: 0, ATTENDANCE_D: 0, ATTENDANCE_G: 0})
        domain = [
            ('employee_id', 'in', employee_ids),
            ('check_in', '>=', fields.Datetime.to_datetime(date_from)),
            ('check_in', '<=', fields.Datetime.to_datetime(date_to).replace(
                hour=23, minute=59, second=59)),
        ]
        for att in self.env['hr.attendance'].sudo().search(domain):
            val = att[ATTENDANCE_FIELD]
            if val in (ATTENDANCE_X, ATTENDANCE_D, ATTENDANCE_G):
                counts[att.employee_id.id][val] += 1
        return counts

    def action_export(self):
        self.ensure_one()

        if self.date_from > self.date_to:
            raise UserError(_("'From' date must be earlier than or equal to 'To' date."))

        columns = self.column_ids.sorted('sequence')
        if not columns:
            raise UserError(_("Please add at least one column before exporting."))

        domain = [
            ('date_from', '>=', self.date_from),
            ('date_to',   '<=', self.date_to),
        ]
        dept_ids = self._get_department_ids()
        if dept_ids is not None:
            domain.append(('department_id', 'in', dept_ids))

        payslips = self.env['hr.payslip'].sudo().search(
            domain, order='department_id, employee_id, date_from'
        )
        if not payslips:
            raise UserError(_("No payslips found for the selected criteria."))

        needs_att = any(
            c.virtual_key in (VIRTUAL_ATT_X, VIRTUAL_ATT_D, VIRTUAL_ATT_G)
            for c in columns
        )
        att_counts = {}
        if needs_att:
            att_counts = self._get_attendance_counts(
                payslips.mapped('employee_id').ids,
                self.date_from, self.date_to,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payslips'

        total_cols = len(columns)

        title_text = 'სახელფასო ფურცლები  %s – %s' % (
            self.date_from.strftime('%d.%m.%Y'),
            self.date_to.strftime('%d.%m.%Y'),
        )
        if self.department_id:
            title_text += '  |  %s' % self.department_id.name

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN
        ws.row_dimensions[1].height = 24

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col.label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col.col_width or 14

        ws.row_dimensions[2].height = 30
        ws.freeze_panes = 'A3'

        col_totals = defaultdict(float)

        for row_idx, slip in enumerate(payslips, start=3):
            row_fill = ALT_FILL if row_idx % 2 == 0 else None

            for col_idx, col in enumerate(columns, start=1):
                value = col._get_cell_value(slip, att_counts)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if row_fill:
                    cell.fill = row_fill

                if col._is_numeric():
                    cell.alignment = RIGHT_ALIGN
                    if col._is_money():
                        cell.number_format = '#,##0.00'
                    if isinstance(value, (int, float)):
                        col_totals[col_idx] += value
                else:
                    cell.alignment = LEFT_ALIGN

        total_row = len(payslips) + 3
        for col_idx, col in enumerate(columns, start=1):
            if col_idx == 1:
                val = 'სულ'
            elif col._is_numeric() and col_idx in col_totals:
                val = col_totals[col_idx]
            else:
                val = ''
            cell = ws.cell(row=total_row, column=col_idx, value=val)
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            if col._is_numeric():
                cell.alignment = RIGHT_ALIGN
                if col._is_money():
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = LEFT_ALIGN
        ws.row_dimensions[total_row].height = 18

        ws.auto_filter.ref = 'A2:%s%d' % (
            get_column_letter(total_cols), total_row - 1
        )

        stream = BytesIO()
        wb.save(stream)
        excel_data = base64.b64encode(stream.getvalue()).decode('utf-8')

        filename = 'payslips_%s_%s.xlsx' % (
            self.date_from.strftime('%d.%m.%Y'),
            self.date_to.strftime('%d.%m.%Y'),
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'datas': excel_data,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })

        config = self.env['payslip.export.config'].sudo().get_config()
        config.save_columns(self.column_ids)

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'self',
        }
