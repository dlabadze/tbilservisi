# -*- coding: utf-8 -*-
from odoo import api, fields, models
from odoo.exceptions import UserError
import base64
from io import BytesIO
from datetime import datetime, date as date_cls


class FleetFuelUploadWizard(models.TransientModel):
	_name = 'fleet.fuel.upload.wizard'
	_description = 'Fleet Fuel Upload Wizard'

	date = fields.Date(
		string='თარიღი',
		required=True,
		default=fields.Date.context_today,
	)

	fuel_type = fields.Selection(
		selection=[
			('gulf', 'გალფი'),
			('rompetrol', 'რომპეტროლი'),
		],
		string='საწვავის ტიპი',
		required=True,
	)

	vendor_id = fields.Many2one(
		comodel_name='res.partner',
		string='მომწოდებელი',
		required=True,
	)

	file_data = fields.Binary(string='Excel ფაილი', required=True)
	file_name = fields.Char(string='ფაილის სახელი')

	def _normalize_license(self, value: str) -> str:
		if not value:
			return ''
		return ''.join(ch for ch in value.upper().strip() if ch.isalnum())

	def _parse_excel_date(self, value, datemode=None):
		"""Try to convert excel cell value to python date.
		- If datetime/date: return date
		- If float/int with datemode: xlrd serial to date
		- If string: try slicing YYYY-MM-DD from 'YYYY-MM-DD HH:MM:SS' then common formats
		Return None if cannot parse.
		"""
		if value is None or value == '':
			return None
		# Already date/datetime
		if isinstance(value, date_cls):
			return value
		if isinstance(value, datetime):
			return value.date()
		# Excel serial number (mostly in xls)
		try:
			from numbers import Number
			is_number = isinstance(value, Number)
		except Exception:
			is_number = False
		if is_number and datemode is not None:
			try:
				import xlrd  # type: ignore
				dt = xlrd.xldate_as_datetime(value, datemode)
				return dt.date()
			except Exception:
				pass
		# String parsing
		if isinstance(value, str):
			text = value.strip()
			# Slice YYYY-MM-DD if text like '2025-09-30 23:39:46'
			if len(text) >= 10 and text[0:4].isdigit() and text[5:7].isdigit() and text[8:10].isdigit():
				try:
					year = int(text[0:4])
					month = int(text[5:7])
					day = int(text[8:10])
					return date_cls(year, month, day)
				except Exception:
					pass
			for fmt in (
				"%Y-%m-%d",
				"%Y-%m-%d %H:%M:%S",
				"%d/%m/%Y",
				"%d-%m-%Y",
				"%d.%m.%Y",
				"%m/%d/%Y",
				"%d/%m/%y",
				"%m/%d/%y",
			):
				try:
					return datetime.strptime(text, fmt).date()
				except Exception:
					continue
		return None

	def _read_excel_rompetrol(self, data: bytes) -> list:
		"""Return list of dicts with keys: date, plate, fuel_text, quantity, station, row_num.
		Columns (1-based, as provided):
		1=date, 4=plate, 5=fuel_text, 8=quantity, 9=station
		"""
		rows = []
		stream = BytesIO(data)
		# Try openpyxl (xlsx) first, then xlrd (xls)
		try:
			import openpyxl  # type: ignore
			wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
			ws = wb.active
			for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
				# Skip header (assume first row is header)
				if idx == 1:
					continue
				date_v = row[0] if len(row) > 0 else None
				plate_v = row[3] if len(row) > 3 else None
				fuel_v = row[4] if len(row) > 4 else None
				qty_v = row[7] if len(row) > 7 else None
				station_v = row[8] if len(row) > 8 else None
				parsed_date = self._parse_excel_date(date_v)
				if plate_v:
					rows.append({
						'row_num': idx,
						'date': parsed_date,
						'plate': str(plate_v).strip(),
						'fuel_text': (str(fuel_v).strip() if fuel_v is not None else ''),
						'quantity': float(qty_v) if qty_v not in (None, '') else 0.0,
						'station': (str(station_v).strip() if station_v is not None else ''),
					})
			return rows
		except Exception:
			# Fallback to xlrd for xls
			stream.seek(0)
			try:
				import xlrd  # type: ignore
				book = xlrd.open_workbook(file_contents=stream.read())
				sheet = book.sheet_by_index(0)
				dm = book.datemode
				for rx in range(1, sheet.nrows):  # skip header
					row = sheet.row(rx)
					def cell_val(ci):
						return row[ci].value if ci < len(row) else None
					date_v = cell_val(0)
					plate_v = cell_val(3)
					fuel_v = cell_val(4)
					qty_v = cell_val(7)
					station_v = cell_val(8)
					parsed_date = self._parse_excel_date(date_v, datemode=dm)
					if plate_v:
						rows.append({
							'row_num': rx + 1,
							'date': parsed_date,
							'plate': str(plate_v).strip(),
							'fuel_text': (str(fuel_v).strip() if fuel_v is not None else ''),
							'quantity': float(qty_v) if qty_v not in (None, '') else 0.0,
							'station': (str(station_v).strip() if station_v is not None else ''),
						})
				return rows
			except Exception as e:  # pragma: no cover
				raise UserError('ვერ წავიკითხე Excel ფაილი. გთხოვთ ატვირთოთ .xlsx ან .xls.\n%s' % e)

	def _read_excel_gulf(self, data: bytes) -> list:
		"""Return list of dicts with keys: date, plate, fuel_text, quantity, station, row_num.
		Gulf mapping (1-based): 1=date, 2=fuel_text, 3=station, 5=plate, 6=quantity
		"""
		rows = []
		stream = BytesIO(data)
		try:
			import openpyxl  # type: ignore
			wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
			ws = wb.active
			for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
				if idx == 1:
					continue
				date_v = row[0] if len(row) > 0 else None
				fuel_v = row[1] if len(row) > 1 else None
				station_v = row[2] if len(row) > 2 else None
				plate_v = row[4] if len(row) > 4 else None
				qty_v = row[5] if len(row) > 5 else None
				parsed_date = self._parse_excel_date(date_v)
				if plate_v:
					rows.append({
						'row_num': idx,
						'date': parsed_date,
						'plate': str(plate_v).strip(),
						'fuel_text': (str(fuel_v).strip() if fuel_v is not None else ''),
						'quantity': float(qty_v) if qty_v not in (None, '') else 0.0,
						'station': (str(station_v).strip() if station_v is not None else ''),
					})
			return rows
		except Exception:
			stream.seek(0)
			try:
				import xlrd  # type: ignore
				book = xlrd.open_workbook(file_contents=stream.read())
				sheet = book.sheet_by_index(0)
				dm = book.datemode
				for rx in range(1, sheet.nrows):
					row = sheet.row(rx)
					def cell_val(ci):
						return row[ci].value if ci < len(row) else None
					date_v = cell_val(0)
					fuel_v = cell_val(1)
					station_v = cell_val(2)
					plate_v = cell_val(4)
					qty_v = cell_val(5)
					parsed_date = self._parse_excel_date(date_v, datemode=dm)
					if plate_v:
						rows.append({
							'row_num': rx + 1,
							'date': parsed_date,
							'plate': str(plate_v).strip(),
							'fuel_text': (str(fuel_v).strip() if fuel_v is not None else ''),
							'quantity': float(qty_v) if qty_v not in (None, '') else 0.0,
							'station': (str(station_v).strip() if station_v is not None else ''),
						})
				return rows
			except Exception as e:  # pragma: no cover
				raise UserError('ვერ წავიკითხე Excel ფაილი. გთხოვთ ატვირთოთ .xlsx ან .xls.\n%s' % e)

	def action_confirm(self):
		self.ensure_one()
		if not self.file_data:
			raise UserError('ატვირთეთ Excel ფაილი.')

		data = base64.b64decode(self.file_data)
		if self.fuel_type == 'rompetrol':
			rows = self._read_excel_rompetrol(data)
		elif self.fuel_type == 'gulf':
			rows = self._read_excel_gulf(data)
		else:
			raise UserError('ამჟამად ხელმისაწვდომია მხოლოდ რომპეტროლი და გალფი.')

		if not rows:
			raise UserError('ფაილში მონაცემები ვერ მოიძებნა.')

		# Build vehicle map by normalized custom field x_studio_carnumbexp
		vehicles = self.env['fleet.vehicle'].search([])
		norm_to_vehicle = {}
		for v in vehicles:
			raw_num = getattr(v, 'x_studio_carnumbexp', False)
			norm = self._normalize_license(raw_num)
			if norm and norm not in norm_to_vehicle:
				norm_to_vehicle[norm] = v

		# Group rows by plate and create one log per vehicle, with multiple lines
		grouped = {}
		raw_plate_by_norm = {}
		for r in rows:
			plate_raw = r.get('plate')
			plate_norm = self._normalize_license(plate_raw)
			if not plate_norm:
				continue
			grouped.setdefault(plate_norm, []).append(r)
			if plate_norm not in raw_plate_by_norm:
				raw_plate_by_norm[plate_norm] = plate_raw

		# Enforce that all plates exist in fleet; abort if any is missing
		missing_norms = [p for p in grouped.keys() if p not in norm_to_vehicle]
		if missing_norms:
			missing_raw = raw_plate_by_norm.get(missing_norms[0], '')
			raise UserError('სატრანსპორტო საშუალება სახელმწიფო ნომრით %s ვერ მოიძებნა.' % (missing_raw))

		# Validate all line dates are present (Excel first column must be parsable)
		for plate_norm, plate_rows in grouped.items():
			for r in plate_rows:
				if not r.get('date'):
					raise UserError('Excel-ის თარიღი (1-ლი სვეტი) არასწორია: რიგი %s, სანომრო %s' % (r.get('row_num') or '', r.get('plate') or ''))

		# Pre-check: if any log exists on wizard date for any matched vehicle, abort
		matched_vehicle_ids = [norm_to_vehicle[p].id for p in grouped.keys()]
		if matched_vehicle_ids:
			existing = self.env['fleet.vehicle.log.fuel'].search([
				('date', '=', self.date),
				('vehicle_id', 'in', matched_vehicle_ids),
			], limit=1)
			if existing:
				plate = existing.license_plate or existing.vehicle_id.license_plate or ''
				raise UserError('სატრანსპორტო საშუალებაზე სახელმწიფო ნომრით %s უკვე გატარებულია საწვავის ჩასხმის ოპერაცია.' % (plate))

		# Resolve Studio service type if field exists
		log_model = self.env['fleet.vehicle.log.fuel']
		studio_field = log_model._fields.get('x_studio_service_type_id')
		studio_service = None
		if studio_field and getattr(studio_field, 'comodel_name', None) == 'fleet.service.type':
			studio_service = self.env['fleet.service.type'].search([('name', '=', 'საწვავის ჩასხმა')], limit=1)

		to_create = []
		for plate_norm, plate_rows in grouped.items():
			vehicle = norm_to_vehicle[plate_norm]
			# Prepare line values from all rows of this vehicle
			line_vals = []
			for r in plate_rows:
				line_vals.append((0, 0, {
					'date': r.get('date'),
					'station': r.get('station') or '',
					'quantity': r.get('quantity') or 0.0,
				}))
			vals = {
				'date': self.date,  # parent uses wizard date
				'vendor_id': self.vendor_id.id,
				'fuel_type_text': plate_rows[0].get('fuel_text') or '',
				'vehicle_id': vehicle.id,
				'line_ids': line_vals,
			}
			# Set Studio service type if available and found
			if studio_field and studio_service:
				vals['x_studio_service_type_id'] = studio_service.id
			to_create.append(vals)

		if not to_create:
			raise UserError('არც ერთი ჩანაწერი არ შეიქმნა. სწორიაო საავტომობილო ნომრები?')

		created = self.env['fleet.vehicle.log.fuel'].create(to_create)
		# Move to Running then Done to finalize properly
		created.button_running()
		created.button_done()
		# Return the main list action to refresh the page after import
		action = self.env.ref('fleet_vehicle_log_fuel.fleet_vehicle_log_fuel_action', raise_if_not_found=False)
		if action:
			return action.read()[0]
		return {'type': 'ir.actions.act_window_close'}
